from fastapi import FastAPI, APIRouter, HTTPException, Query, Request, Response, Cookie, UploadFile, File, Body
from fastapi.responses import StreamingResponse, FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.gzip import GZipMiddleware
from supabase import create_client, Client
import os
import logging
import asyncio
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from difflib import SequenceMatcher
import re
import io
import httpx
import json
import shutil
from passlib.context import CryptContext
from functools import lru_cache
import time as _time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import resend
import google.generativeai as genai
import backup_service

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Supabase connection
SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_KEY')
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Resend Email configuration
RESEND_API_KEY = os.environ.get('RESEND_API_KEY')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'onboarding@resend.dev')
resend.api_key = RESEND_API_KEY

# Google Gemini API configuration
GOOGLE_API_KEY = os.environ.get('GOOGLE_API_KEY')
if GOOGLE_API_KEY:
    genai.configure(api_key=GOOGLE_API_KEY)

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Uploads directory
UPLOADS_DIR = ROOT_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)

# ============ ACTIVITY LOG HELPER ============

async def log_activity(
    activity_type: str,
    title: str,
    subtitle: str = "",
    customer_id: str = None,
    customer_name: str = "",
    user_email: str = "",
    user_name: str = "",
    metadata: dict = None
):
    """Log an activity to the activity_log table"""
    try:
        activity = {
            "id": str(uuid.uuid4()),
            "activity_type": activity_type,
            "title": title,
            "subtitle": subtitle,
            "customer_id": customer_id,
            "customer_name": customer_name,
            "user_email": user_email,
            "user_name": user_name,
            "metadata": metadata or {},
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        supabase.table("activity_log").insert(activity).execute()
        # Invalidate activity feed cache so new entries appear immediately
        try:
            _activity_cache.clear()
        except Exception:
            pass
    except Exception as e:
        logging.error(f"Activity log error: {e}")

# ============ MODELS ============

class ContactPerson(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    role: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    is_primary: Optional[bool] = False

class ContactInfo(BaseModel):
    contact_person: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""

class CustomerBase(BaseModel):
    company_name: Optional[str] = ""
    market: Optional[str] = ""
    application: Optional[str] = ""
    city: Optional[str] = ""
    district: Optional[str] = ""
    website: Optional[str] = ""
    status: Optional[str] = "Beklemede"
    contact_info: Optional[ContactInfo] = None
    contacts: Optional[List[ContactPerson]] = []
    potential_value: Optional[float] = 0
    next_followup_date: Optional[str] = ""
    assigned_to: Optional[str] = ""
    competitor: Optional[str] = ""
    partner: Optional[str] = ""
    potential_level: Optional[str] = "Düşük"
    products: Optional[List[str]] = []
    description: Optional[str] = ""
    notes: Optional[str] = ""
    notes_list: Optional[List[dict]] = []
    documents: Optional[List[dict]] = []
    tags: Optional[List[str]] = []
    is_followup: Optional[bool] = False

class CustomerCreate(CustomerBase):
    pass

class CustomerUpdate(CustomerBase):
    company_name: Optional[str] = None

class Customer(CustomerBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class VisitBase(BaseModel):
    customer_id: Optional[str] = ""
    visit_date: Optional[str] = ""
    visit_type: Optional[str] = "Yüz Yüze"
    visited_by: Optional[str] = ""  # Ziyaret eden kişi
    contact_person: Optional[str] = ""
    notes: Optional[str] = ""
    next_visit_date: Optional[str] = ""
    outcome: Optional[str] = ""
    is_followup: Optional[bool] = False

class VisitCreate(VisitBase):
    pass

class Visit(VisitBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class DynamicOption(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    field_name: str
    value: str
    color: Optional[str] = None

class SimilarityResult(BaseModel):
    customer_id: str
    company_name: str
    similarity_score: float
    match_type: str

# Call Model
class CallBase(BaseModel):
    customer_id: str
    call_date: Optional[str] = ""
    caller_name: Optional[str] = ""
    contact_person: Optional[str] = ""
    phone_number: Optional[str] = ""
    duration_minutes: Optional[int] = 0
    call_type: Optional[str] = "Giden"
    outcome: Optional[str] = ""
    notes: Optional[str] = ""
    next_action: Optional[str] = ""
    next_action_date: Optional[str] = ""

class CallCreate(CallBase):
    pass

class Call(CallBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Saved Filter Model
class FilterCondition(BaseModel):
    field: str
    operator: str
    value: Optional[str] = ""

class SavedFilterBase(BaseModel):
    name: str
    conditions: List[FilterCondition] = []
    logic: str = "AND"

class SavedFilterCreate(SavedFilterBase):
    pass

class SavedFilter(SavedFilterBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_by: Optional[str] = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Kanban View Model
class KanbanViewBase(BaseModel):
    name: str
    group_by: str
    description: Optional[str] = ""

class KanbanViewCreate(KanbanViewBase):
    pass

class KanbanView(KanbanViewBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_by: Optional[str] = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ============ HELPER FUNCTIONS ============

def normalize_text(text: str) -> str:
    if not text:
        return ""
    text = text.lower().strip()
    text = re.sub(r'[^\w\s]', '', text)
    text = re.sub(r'\s+', ' ', text)
    return text

def calculate_similarity(str1: str, str2: str) -> float:
    if not str1 or not str2:
        return 0.0
    norm1 = normalize_text(str1)
    norm2 = normalize_text(str2)
    if not norm1 or not norm2:
        return 0.0
    if norm1 == norm2:
        return 1.0
    # Substring containment boost — "Lamtek" should strongly match "Lamtek Makina"
    if norm1 in norm2 or norm2 in norm1:
        shorter = min(len(norm1), len(norm2))
        longer = max(len(norm1), len(norm2))
        ratio = shorter / longer if longer else 0
        return max(ratio, 0.85)
    return SequenceMatcher(None, norm1, norm2).ratio()

def normalize_phone(phone: str) -> str:
    if not phone:
        return ""
    return re.sub(r'[^\d]', '', phone)

def normalize_website(url: str) -> str:
    if not url:
        return ""
    url = url.lower().strip()
    url = re.sub(r'^https?://', '', url)
    url = re.sub(r'^www\.', '', url)
    url = re.sub(r'/$', '', url)
    return url

def find_similar_customers(customer_data: dict, exclude_id: str = None) -> List[SimilarityResult]:
    results = []
    
    # Paginate to bypass Supabase 1000 row limit — DB may have >1000 customers
    existing_customers = fetch_all_rows(
        "customers", "id, company_name, website, contact_info"
    )
    if exclude_id:
        existing_customers = [c for c in existing_customers if c.get("id") != exclude_id]
    
    for existing in existing_customers:
        max_similarity = 0.0
        match_type = ""
        
        name_sim = calculate_similarity(
            customer_data.get("company_name", ""),
            existing.get("company_name", "")
        )
        if name_sim > max_similarity:
            max_similarity = name_sim
            match_type = "Firma Adı"
        
        if customer_data.get("contact_info") and existing.get("contact_info"):
            phone1 = normalize_phone(customer_data.get("contact_info", {}).get("phone", ""))
            phone2 = normalize_phone(existing.get("contact_info", {}).get("phone", ""))
            if phone1 and phone2:
                if phone1 == phone2:
                    max_similarity = 1.0
                    match_type = "Telefon"
                elif phone1 in phone2 or phone2 in phone1:
                    if 0.9 > max_similarity:
                        max_similarity = 0.9
                        match_type = "Telefon"
        
        web1 = normalize_website(customer_data.get("website", ""))
        web2 = normalize_website(existing.get("website", ""))
        if web1 and web2:
            if web1 == web2:
                max_similarity = 1.0
                match_type = "Web Sitesi"
            else:
                web_sim = calculate_similarity(web1, web2)
                if web_sim > max_similarity:
                    max_similarity = web_sim
                    match_type = "Web Sitesi"
        
        if max_similarity >= 0.7:
            results.append(SimilarityResult(
                customer_id=existing["id"],
                company_name=existing["company_name"],
                similarity_score=round(max_similarity * 100, 1),
                match_type=match_type
            ))
    
    results.sort(key=lambda x: x.similarity_score, reverse=True)
    return results[:5]

# ============ CUSTOMER ENDPOINTS ============

@api_router.get("/")
async def root():
    return {"message": "CRM API is running"}

@api_router.post("/customers", response_model=Customer)
async def create_customer(customer: CustomerCreate):
    customer_dict = customer.model_dump()
    if customer_dict.get("contact_info"):
        customer_dict["contact_info"] = dict(customer_dict["contact_info"])
    
    customer_obj = Customer(**customer_dict)
    doc = customer_obj.model_dump()
    
    # Convert lists to proper format for PostgreSQL
    doc["products"] = doc.get("products", [])
    doc["tags"] = doc.get("tags", [])
    
    response = supabase.table("customers").insert(doc).execute()
    
    # Invalidate caches
    _filter_options_cache["data"] = None
    _stats_cache["data"] = None
    _invalidate_kanban_cache()
    # Log activity
    await log_activity(
        activity_type="customer_created",
        title=f"Yeni müşteri: {customer_obj.company_name}",
        subtitle=f"Durum: {customer_obj.status or 'Beklemede'}",
        customer_id=customer_obj.id,
        customer_name=customer_obj.company_name
    )
    
    return customer_obj

@api_router.get("/customers")
async def get_customers(
    search: Optional[str] = None,
    market: Optional[str] = None,
    application: Optional[str] = None,
    city: Optional[str] = None,
    district: Optional[str] = None,
    status: Optional[str] = None,
    competitor: Optional[str] = None,
    partner: Optional[str] = None,
    assigned_to: Optional[str] = None,
    is_followup: Optional[bool] = None,
    page: Optional[int] = 1,
    limit: Optional[int] = 50,
    sort_by: Optional[str] = "created_at",
    sort_order: Optional[str] = "desc"
):
    """Get customers with server-side pagination, sorting and filtering.

    Performance: this uses a SINGLE Supabase query with count="exact" and
    .range() — previously the code issued TWO round-trips (one to count,
    downloading every matching row, and one to paginate). On 3600+ customers
    that wasted ~400ms per request.
    """
    try:
        # Whitelist of allowed sort fields. "data_score" is a computed field
        # (count of filled fields per customer) — handled separately because
        # Supabase can't order by a derived value.
        allowed_sort = {"created_at", "updated_at", "company_name", "city", "market", "status", "data_score"}
        sort_field = sort_by if sort_by in allowed_sort else "created_at"
        is_desc = (sort_order or "desc").lower() == "desc"

        def _apply_filters(q):
            if market: q = q.eq("market", market)
            if application: q = q.eq("application", application)
            if city: q = q.eq("city", city)
            if district: q = q.eq("district", district)
            if status: q = q.eq("status", status)
            if competitor: q = q.eq("competitor", competitor)
            if partner: q = q.eq("partner", partner)
            if is_followup is not None: q = q.eq("is_followup", is_followup)
            if assigned_to: q = q.ilike("assigned_to", f"%{assigned_to}%")
            if search:
                search_term = search.strip()
                if search_term:
                    escaped = search_term.replace(",", " ").replace("(", " ").replace(")", " ")
                    pattern = f"*{escaped}*"
                    # Search across text columns only. Array/JSON columns (tags,
                    # products, notes_list) are matched client-side via the
                    # _matchInfo helper — Supabase's `::text` cast inside or_()
                    # was returning 0 rows even when total>0 (silent failure).
                    or_filter = ",".join([
                        f"company_name.ilike.{pattern}",
                        f"market.ilike.{pattern}",
                        f"application.ilike.{pattern}",
                        f"city.ilike.{pattern}",
                        f"district.ilike.{pattern}",
                        f"partner.ilike.{pattern}",
                        f"competitor.ilike.{pattern}",
                        f"assigned_to.ilike.{pattern}",
                        f"status.ilike.{pattern}",
                        f"potential_level.ilike.{pattern}",
                        f"description.ilike.{pattern}",
                        f"notes.ilike.{pattern}",
                        f"website.ilike.{pattern}",
                    ])
                    q = q.or_(or_filter)
            return q

        offset = (page - 1) * limit

        # ----- Special path: sort by computed "data_score" -----
        # Strategy: maintain a 60s cache of {id -> score} computed across ALL
        # customers (light columns). For each request, fetch only matching
        # IDs (cheap), look up scores from cache, sort, paginate, then fetch
        # the page's full rows in one query.
        if sort_field == "data_score":
            score_cols = [
                "company_name", "market", "application", "city", "district",
                "website", "contact_info", "contacts", "potential_value",
                "next_followup_date", "assigned_to", "competitor", "partner",
                "potential_level", "products", "description", "notes",
                "notes_list", "documents", "tags",
            ]

            def _is_filled(v):
                if v is None: return False
                if isinstance(v, str): return bool(v.strip())
                if isinstance(v, (list, dict)): return len(v) > 0
                if isinstance(v, (int, float)): return v != 0
                return bool(v)

            now_ts = _time.time()
            score_map = _data_score_cache.get("map")
            if not score_map or (now_ts - _data_score_cache["ts"]) > _DATA_SCORE_CACHE_TTL:
                # Recompute for ALL customers
                select_str = "id, " + ", ".join(score_cols)
                all_rows = []
                chunk_offset = 0
                chunk_size = 1000
                while True:
                    resp = supabase.table("customers").select(select_str).range(
                        chunk_offset, chunk_offset + chunk_size - 1
                    ).execute()
                    rows = resp.data or []
                    all_rows.extend(rows)
                    if len(rows) < chunk_size:
                        break
                    chunk_offset += chunk_size
                score_map = {
                    r["id"]: sum(1 for c in score_cols if _is_filled(r.get(c)))
                    for r in all_rows
                }
                _data_score_cache["map"] = score_map
                _data_score_cache["ts"] = now_ts

            # Fetch ONLY ids+name for the filtered set (so we can sort).
            # OPTIMIZATION: if there are no filters/search, we already have
            # every customer's score in the cache. Skip the extra round-trip
            # and use the cached map directly. Saves ~500ms on warm calls.
            has_filters = bool(
                market or application or city or district or status or
                competitor or partner or assigned_to or search or
                is_followup is not None
            )
            if not has_filters:
                # Pull only id+company_name for tie-breaker, single chunked fetch
                filtered_rows = []
                chunk_offset = 0
                chunk_size = 1000
                while True:
                    resp = supabase.table("customers").select("id, company_name").range(
                        chunk_offset, chunk_offset + chunk_size - 1
                    ).execute()
                    rows = resp.data or []
                    filtered_rows.extend(rows)
                    if len(rows) < chunk_size:
                        break
                    chunk_offset += chunk_size
                total_count = len(filtered_rows)
            else:
                filtered_rows = []
                total_count = 0
                chunk_offset = 0
                chunk_size = 1000
                first = True
                while True:
                    q_chunk = supabase.table("customers").select(
                        "id, company_name", count="exact" if first else None
                    )
                    q_chunk = _apply_filters(q_chunk)
                    q_chunk = q_chunk.range(chunk_offset, chunk_offset + chunk_size - 1)
                    resp = q_chunk.execute()
                    if first:
                        total_count = resp.count or 0
                        first = False
                    rows = resp.data or []
                    filtered_rows.extend(rows)
                    if len(rows) < chunk_size:
                        break
                    chunk_offset += chunk_size

            # Sort by score (computed) then company_name as tie-breaker
            filtered_rows.sort(
                key=lambda r: (score_map.get(r["id"], 0), r.get("company_name") or ""),
                reverse=is_desc,
            )

            page_rows = filtered_rows[offset:offset + limit]
            page_ids = [r["id"] for r in page_rows]
            if page_ids:
                full_resp = supabase.table("customers").select("*").in_("id", page_ids).execute()
                full_map = {c["id"]: c for c in (full_resp.data or [])}
                customers = []
                for r in page_rows:
                    full = full_map.get(r["id"])
                    if full:
                        full["data_score"] = score_map.get(r["id"], 0)
                        customers.append(full)
            else:
                customers = []

            total_pages = (total_count + limit - 1) // limit if total_count > 0 else 1
            return {
                "data": customers,
                "total": total_count,
                "page": page,
                "limit": limit,
                "total_pages": total_pages,
            }

        # ----- Normal Supabase-native sort path -----
        if limit > 1000:
            customers = []
            page_size = 1000
            current_offset = offset
            remaining = limit
            total_count = 0
            first_chunk = True
            while remaining > 0:
                chunk_size = min(page_size, remaining)
                q = supabase.table("customers").select(
                    "*", count="exact" if first_chunk else None
                )
                q = _apply_filters(q)
                q = q.order(sort_field, desc=is_desc).range(
                    current_offset, current_offset + chunk_size - 1
                )
                chunk_resp = q.execute()
                if first_chunk:
                    total_count = chunk_resp.count or 0
                    first_chunk = False
                chunk_data = chunk_resp.data or []
                customers.extend(chunk_data)
                if len(chunk_data) < chunk_size:
                    break
                current_offset += chunk_size
                remaining -= chunk_size
        else:
            # SINGLE query: data + count in one round-trip
            q = supabase.table("customers").select("*", count="exact")
            q = _apply_filters(q)
            q = q.order(sort_field, desc=is_desc).range(offset, offset + limit - 1)
            response = q.execute()
            customers = response.data or []
            total_count = response.count or 0

        total_pages = (total_count + limit - 1) // limit if total_count > 0 else 1

        return {
            "data": customers,
            "total": total_count,
            "page": page,
            "limit": limit,
            "total_pages": total_pages
        }
    except Exception as e:
        logging.error(f"Error fetching customers: {e}")
        # Fallback to simple query
        response = supabase.table("customers").select("*").order("created_at", desc=True).limit(limit).execute()
        return {
            "data": response.data or [],
            "total": len(response.data) if response.data else 0,
            "page": 1,
            "limit": limit,
            "total_pages": 1
        }

# Simple in-memory cache for filter options
_filter_options_cache = {"data": None, "timestamp": 0}
_FILTER_CACHE_TTL = 60  # 60 seconds cache

# Activity feed cache (per limit). Invalidated whenever log_activity() inserts.
_activity_cache = {}
_ACTIVITY_CACHE_TTL = 15

# Kanban cache (per group_by key)
_kanban_cache = {}
_KANBAN_CACHE_TTL = 30  # 30 seconds cache

# Latest calls per customer cache (30s — invalidated on new call insert)
_latest_calls_cache = {"data": None, "ts": 0}
_LATEST_CALLS_CACHE_TTL = 30

# Distribution cache (per field+followup_only key) — 30s TTL.
_distribution_cache = {}
_DISTRIBUTION_CACHE_TTL = 30

def _invalidate_kanban_cache():
    """Clear all kanban + chart caches (called on any customer write)"""
    _kanban_cache.clear()
    # Also clear caches that share the same underlying customers data
    _distribution_cache.clear()
    _followups_cache["data"] = None
    _data_score_cache["map"] = None


def fetch_all_rows(table: str, select_cols: str = "*", page_size: int = 1000):
    """
    Fetch ALL rows from a Supabase table via pagination.
    Supabase caps single .execute() at 1000 rows by default, so we page through.
    """
    rows = []
    offset = 0
    while True:
        resp = supabase.table(table).select(select_cols).range(offset, offset + page_size - 1).execute()
        chunk = resp.data or []
        rows.extend(chunk)
        if len(chunk) < page_size:
            break
        offset += page_size
    return rows

@api_router.get("/customers/filter-options")
async def get_customer_filter_options():
    """Return unique filter values - cached for 60 seconds"""
    now = _time.time()
    if _filter_options_cache["data"] and (now - _filter_options_cache["timestamp"]) < _FILTER_CACHE_TTL:
        return _filter_options_cache["data"]
    
    fields = ["market", "application", "city", "competitor", "partner", "assigned_to"]
    select_str = ", ".join(fields)
    
    # Single query with high limit
    all_data = []
    page_size = 1000
    offset = 0
    while True:
        response = supabase.table("customers").select(select_str).range(offset, offset + page_size - 1).execute()
        if not response.data:
            break
        all_data.extend(response.data)
        if len(response.data) < page_size:
            break
        offset += page_size
    
    result = {}
    for field in fields:
        result[field] = sorted(set(row[field] for row in all_data if row.get(field)))
    
    _filter_options_cache["data"] = result
    _filter_options_cache["timestamp"] = now
    return result

@api_router.get("/customers/{customer_id}", response_model=Customer)
async def get_customer(customer_id: str):
    response = supabase.table("customers").select("*").eq("id", customer_id).execute()
    if not response.data:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    return response.data[0]

@api_router.get("/calls/latest-per-customer")
async def get_latest_calls_per_customer():
    """Return only the latest call outcome per customer - lightweight endpoint for Customers table.

    Cached for 30 seconds (invalidated when a new call is logged).
    """
    now = _time.time()
    cached = _latest_calls_cache.get("data")
    if cached and (now - _latest_calls_cache["ts"]) < _LATEST_CALLS_CACHE_TTL:
        return cached
    # Paginated fetch (Supabase default cap is 1000 per query). Without this,
    # only the first ~1000 calls are returned and many customers miss their
    # latest outcome.
    all_calls = []
    page_size = 1000
    offset = 0
    while True:
        response = (
            supabase.table("calls")
            .select("customer_id, outcome, created_at")
            .order("created_at", desc=True)
            .range(offset, offset + page_size - 1)
            .execute()
        )
        if not response.data:
            break
        all_calls.extend(response.data)
        if len(response.data) < page_size:
            break
        offset += page_size

    latest = {}
    for call in all_calls:
        cid = call.get("customer_id")
        if cid and cid not in latest:
            latest[cid] = call.get("outcome", "")

    _latest_calls_cache["data"] = latest
    _latest_calls_cache["ts"] = now
    return latest


@api_router.put("/customers/{customer_id}", response_model=Customer)
async def update_customer(customer_id: str, customer: CustomerUpdate):
    response = supabase.table("customers").select("*").eq("id", customer_id).execute()
    if not response.data:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    
    old_customer = response.data[0]
    # Use exclude_unset=True so only explicitly-sent fields are updated.
    # This is essential for inline editing where only 1 field is sent.
    update_data = {k: v for k, v in customer.model_dump(exclude_unset=True).items() if v is not None}
    if update_data.get("contact_info"):
        update_data["contact_info"] = dict(update_data["contact_info"])
    
    # Normalize status to match Kanban columns
    if "status" in update_data and update_data["status"]:
        update_data["status"] = normalize_status(update_data["status"])
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    supabase.table("customers").update(update_data).eq("id", customer_id).execute()
    
    # Invalidate caches
    _filter_options_cache["data"] = None
    _stats_cache["data"] = None
    _invalidate_kanban_cache()
    
    updated = supabase.table("customers").select("*").eq("id", customer_id).execute()
    new_customer = updated.data[0]
    
    # Log specific changes
    company_name = new_customer.get("company_name", "")
    
    # Check status change
    if "status" in update_data and old_customer.get("status") != update_data.get("status"):
        await log_activity(
            activity_type="status_changed",
            title=f"Durum değişti: {company_name}",
            subtitle=f"{old_customer.get('status', 'Yok')} → {update_data.get('status')}",
            customer_id=customer_id,
            customer_name=company_name
        )
    # Check follow-up change
    elif "is_followup" in update_data and old_customer.get("is_followup") != update_data.get("is_followup"):
        status = "Takibe alındı" if update_data.get("is_followup") else "Takipten çıkarıldı"
        await log_activity(
            activity_type="followup_changed",
            title=f"{status}: {company_name}",
            subtitle="",
            customer_id=customer_id,
            customer_name=company_name
        )
    # General update
    elif len(update_data) > 1:  # More than just updated_at
        changed_fields = [k for k in update_data.keys() if k != "updated_at"]
        await log_activity(
            activity_type="customer_updated",
            title=f"Güncellendi: {company_name}",
            subtitle=f"Değişen: {', '.join(changed_fields[:3])}",
            customer_id=customer_id,
            customer_name=company_name
        )
    
    return new_customer

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user_from_request(request, session_token)
    if not check_permission(user, "can_delete"):
        raise HTTPException(status_code=403, detail="Bu işlem için silme yetkisi gerekli")

    # Get customer name before delete
    customer = supabase.table("customers").select("company_name").eq("id", customer_id).execute()
    if not customer.data:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    
    company_name = customer.data[0]["company_name"] if customer.data else "Bilinmiyor"
    
    response = supabase.table("customers").delete().eq("id", customer_id).execute()
    
    # Invalidate caches
    _filter_options_cache["data"] = None
    _stats_cache["data"] = None
    _invalidate_kanban_cache()
    # Log activity
    await log_activity(
        activity_type="customer_deleted",
        title=f"Silindi: {company_name}",
        subtitle="",
        customer_id=None,
        customer_name=company_name,
        user_email=""
    )
    
    return {"message": "Müşteri silindi"}

@api_router.post("/customers/{customer_id}/contacts", response_model=Customer)
async def add_contact_person(customer_id: str, contact: ContactPerson):
    response = supabase.table("customers").select("*").eq("id", customer_id).execute()
    if not response.data:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    
    customer = response.data[0]
    company_name = customer.get("company_name", "")
    contacts = customer.get("contacts") or []
    contacts.append(contact.model_dump())
    
    supabase.table("customers").update({
        "contacts": contacts,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }).eq("id", customer_id).execute()
    
    # Log activity
    await log_activity(
        activity_type="contact_added",
        title=f"Kişi eklendi: {contact.name}",
        subtitle=f"{company_name} - {contact.role or contact.title or ''}",
        customer_id=customer_id,
        customer_name=company_name
    )
    
    updated = supabase.table("customers").select("*").eq("id", customer_id).execute()
    return updated.data[0]

@api_router.delete("/customers/{customer_id}/contacts/{contact_id}")
async def delete_contact_person(customer_id: str, contact_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user_from_request(request, session_token)
    if not check_permission(user, "can_delete"):
        raise HTTPException(status_code=403, detail="Bu işlem için silme yetkisi gerekli")
    response = supabase.table("customers").select("*").eq("id", customer_id).execute()
    if not response.data:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    
    customer = response.data[0]
    company_name = customer.get("company_name", "")
    contacts = customer.get("contacts") or []
    
    # Find contact name before removing
    contact_name = ""
    for c in contacts:
        if c.get("id") == contact_id:
            contact_name = c.get("name", "")
            break
    
    contacts = [c for c in contacts if c.get("id") != contact_id]
    
    supabase.table("customers").update({
        "contacts": contacts,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }).eq("id", customer_id).execute()
    
    # Log activity
    if contact_name:
        await log_activity(
            activity_type="contact_deleted",
            title=f"Kişi silindi: {contact_name}",
            subtitle=company_name,
            customer_id=customer_id,
            customer_name=company_name
        )
    
    return {"message": "Kişi silindi"}

@api_router.post("/customers/check-similar")
async def check_similar_customers(customer_data: dict):
    results = find_similar_customers(customer_data)
    return results

@api_router.post("/import/preview")
async def preview_import(data: dict):
    """Preview import data and check for similar/duplicate customers - optimized"""
    rows = data.get("rows", [])
    
    if not rows:
        return {"items": [], "total_new": 0, "total_similar": 0, "total_duplicate": 0}
    
    # Fetch ALL existing customers ONCE for comparison
    existing_customers = []
    page_size = 1000
    offset = 0
    while True:
        response = supabase.table("customers").select("id, company_name, website, contact_info").range(offset, offset + page_size - 1).execute()
        if not response.data:
            break
        existing_customers.extend(response.data)
        if len(response.data) < page_size:
            break
        offset += page_size
    
    # Pre-normalize existing customer data for fast comparison
    existing_normalized = []
    for ec in existing_customers:
        existing_normalized.append({
            "id": ec["id"],
            "company_name": ec.get("company_name", ""),
            "norm_name": normalize_text(ec.get("company_name", "")),
            "norm_web": normalize_website(ec.get("website", "")),
            "norm_phone": normalize_phone((ec.get("contact_info") or {}).get("phone", ""))
        })
    
    items = []
    total_new = 0
    total_similar = 0
    total_duplicate = 0
    
    for row in rows:
        row_num = row.get("row_num", 0)
        company_name = row.get("company_name", "")
        website = row.get("website", "")
        
        norm_name = normalize_text(company_name)
        norm_web = normalize_website(website)
        
        # Compare against cached existing customers
        similar_results = []
        for ec in existing_normalized:
            max_sim = 0.0
            match_type = ""
            
            # Name comparison
            if norm_name and ec["norm_name"]:
                if norm_name == ec["norm_name"]:
                    name_sim = 1.0
                else:
                    name_sim = SequenceMatcher(None, norm_name, ec["norm_name"]).ratio()
                if name_sim > max_sim:
                    max_sim = name_sim
                    match_type = "Firma Adı"
            
            # Website comparison
            if norm_web and ec["norm_web"]:
                if norm_web == ec["norm_web"]:
                    max_sim = 1.0
                    match_type = "Web Sitesi"
                else:
                    web_sim = SequenceMatcher(None, norm_web, ec["norm_web"]).ratio()
                    if web_sim > max_sim:
                        max_sim = web_sim
                        match_type = "Web Sitesi"
            
            if max_sim >= 0.7:
                similar_results.append({
                    "customer_id": ec["id"],
                    "company_name": ec["company_name"],
                    "similarity_score": round(max_sim * 100, 1),
                    "match_type": match_type
                })
        
        # Sort by score and take top 5
        similar_results.sort(key=lambda x: x["similarity_score"], reverse=True)
        similar_results = similar_results[:5]
        
        if similar_results:
            max_score = similar_results[0]["similarity_score"]
            if max_score >= 95:
                status = "duplicate"
                total_duplicate += 1
            else:
                status = "similar"
                total_similar += 1
        else:
            status = "new"
            total_new += 1
        
        items.append({
            "row_num": row_num,
            "company_name": company_name,
            "similarity_status": status,
            "similar_customers": similar_results
        })
    
    return {
        "items": items,
        "total_new": total_new,
        "total_similar": total_similar,
        "total_duplicate": total_duplicate
    }

@api_router.post("/customers/{customer_id}/merge")
async def merge_customer(customer_id: str, new_data: dict):
    """Merge import data with existing customer - only fill empty fields"""
    # Get existing customer
    response = supabase.table("customers").select("*").eq("id", customer_id).execute()
    if not response.data:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    
    existing = response.data[0]
    update_data = {}
    
    # Fields to potentially merge
    merge_fields = [
        "market", "application", "city", "district", "website", 
        "competitor", "partner", "potential_level", "assigned_to",
        "notes", "description"
    ]
    
    for field in merge_fields:
        existing_value = existing.get(field, "")
        new_value = new_data.get(field, "")
        # Only update if existing is empty and new has value
        if (not existing_value or existing_value == "") and new_value:
            update_data[field] = new_value
    
    # Handle contact_info specially
    existing_contact = existing.get("contact_info", {}) or {}
    new_contact = new_data.get("contact_info", {}) or {}
    merged_contact = {**existing_contact}
    
    for key in ["contact_person", "email", "phone"]:
        if (not existing_contact.get(key)) and new_contact.get(key):
            merged_contact[key] = new_contact[key]
    
    if merged_contact != existing_contact:
        update_data["contact_info"] = merged_contact
    
    # Handle products - merge arrays
    existing_products = existing.get("products", []) or []
    new_products = new_data.get("products", []) or []
    if new_products:
        merged_products = list(set(existing_products + new_products))
        if merged_products != existing_products:
            update_data["products"] = merged_products
    
    # Handle tags - merge arrays
    existing_tags = existing.get("tags", []) or []
    new_tags = new_data.get("tags", []) or []
    if new_tags:
        merged_tags = list(set(existing_tags + new_tags))
        if merged_tags != existing_tags:
            update_data["tags"] = merged_tags
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        supabase.table("customers").update(update_data).eq("id", customer_id).execute()
        
        # Log activity
        await log_activity(
            activity_type="customer_updated",
            title=f"Müşteri güncellendi (Import): {existing.get('company_name')}",
            subtitle=f"{len(update_data)} alan güncellendi",
            customer_id=customer_id,
            customer_name=existing.get("company_name")
        )
    
    # Return updated customer
    updated = supabase.table("customers").select("*").eq("id", customer_id).execute()
    return {
        "message": f"{len(update_data)} alan güncellendi" if update_data else "Güncellenecek alan yok",
        "updated_fields": list(update_data.keys()),
        "customer": updated.data[0] if updated.data else None
    }


# ============ DUPLICATE DETECTION & MERGE ============

# Cache duplicates result for 5 min (full scan is expensive)
_duplicates_cache = {"data": None, "ts": 0, "threshold": None}
_DUPLICATES_CACHE_TTL = 300


@api_router.get("/customers-duplicates")
async def find_duplicate_customers(
    threshold: int = Query(90, ge=70, le=100, description="Minimum similarity % (70-100)"),
    request: Request = None,
    session_token: Optional[str] = Cookie(None),
):
    """Scan ALL customers and group ones likely to be duplicates.

    Uses fuzzy name match + context (city / website). Returns groups of 2+
    customers each with a confidence score. Heavy, so cached for 5 minutes
    per threshold.
    """
    user = await get_current_user_from_request(request, session_token)
    if not check_admin_permission(user):
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")

    now_ts = _time.time()
    if (
        _duplicates_cache["data"] is not None
        and _duplicates_cache["threshold"] == threshold
        and (now_ts - _duplicates_cache["ts"]) < _DUPLICATES_CACHE_TTL
    ):
        return _duplicates_cache["data"]

    # Fetch lightweight customer rows (chunked for >1000)
    rows = []
    chunk_offset = 0
    chunk_size = 1000
    while True:
        resp = supabase.table("customers").select(
            "id, company_name, website, city, district, market, contact_info, "
            "partner, competitor, application, status, created_at"
        ).range(chunk_offset, chunk_offset + chunk_size - 1).execute()
        chunk = resp.data or []
        rows.extend(chunk)
        if len(chunk) < chunk_size:
            break
        chunk_offset += chunk_size

    # Pre-normalize once
    normed = []
    for r in rows:
        normed.append({
            "id": r["id"],
            "row": r,
            "name_norm": normalize_text(r.get("company_name") or ""),
            "web_norm": normalize_website(r.get("website") or ""),
            "city_norm": normalize_text(r.get("city") or ""),
            "phone_norm": normalize_phone(
                (r.get("contact_info") or {}).get("phone") or ""
                if isinstance(r.get("contact_info"), dict) else ""
            ),
        })

    # Union-Find for grouping
    parent = {r["id"]: r["id"] for r in normed}
    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x
    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    pair_meta = {}  # (id_a, id_b) -> {score, match_type}
    thresh = threshold / 100.0

    # Pre-bucket by first 3 chars of normalized name for O(n²) -> ~O(n * b)
    buckets = {}
    for n in normed:
        if not n["name_norm"]:
            continue
        key = n["name_norm"][:3]
        buckets.setdefault(key, []).append(n)

    # Compare within bucket + neighboring buckets (in case of slight prefix diff)
    bucket_keys = sorted(buckets.keys())
    for i, key in enumerate(bucket_keys):
        candidates = buckets[key][:]
        # Also compare with next bucket (handles "hak" vs "hak ")
        if i + 1 < len(bucket_keys) and abs(ord(bucket_keys[i + 1][0]) - ord(key[0])) <= 1:
            candidates.extend(buckets[bucket_keys[i + 1]])

        m = len(candidates)
        for a_idx in range(m):
            a = candidates[a_idx]
            for b_idx in range(a_idx + 1, m):
                b = candidates[b_idx]
                if a["id"] == b["id"]:
                    continue
                # Quick reject if name length differs too much
                la, lb = len(a["name_norm"]), len(b["name_norm"])
                if la == 0 or lb == 0:
                    continue
                if abs(la - lb) / max(la, lb) > (1 - thresh + 0.15):
                    continue

                name_sim = SequenceMatcher(None, a["name_norm"], b["name_norm"]).ratio()

                # Boost confidence with context matches
                same_phone = a["phone_norm"] and b["phone_norm"] and a["phone_norm"] == b["phone_norm"]
                same_web = a["web_norm"] and b["web_norm"] and a["web_norm"] == b["web_norm"]
                same_city = a["city_norm"] and b["city_norm"] and a["city_norm"] == b["city_norm"]

                if same_phone or same_web:
                    final_score = max(name_sim, 0.95)
                    match_type = "Telefon" if same_phone else "Web Sitesi"
                elif name_sim >= thresh:
                    if same_city:
                        final_score = min(1.0, name_sim + 0.05)
                        match_type = "İsim + Şehir"
                    else:
                        final_score = name_sim
                        match_type = "İsim"
                else:
                    continue

                if final_score >= thresh:
                    pair_meta[(a["id"], b["id"])] = {
                        "score": round(final_score * 100, 1),
                        "match_type": match_type,
                    }
                    union(a["id"], b["id"])

    # Build groups from union-find roots
    by_root = {}
    pair_index = {}  # id -> list of (other_id, meta)
    for (id_a, id_b), meta in pair_meta.items():
        pair_index.setdefault(id_a, []).append((id_b, meta))
        pair_index.setdefault(id_b, []).append((id_a, meta))

    in_pair = set()
    for k in pair_meta.keys():
        in_pair.add(k[0]); in_pair.add(k[1])

    id_to_row = {n["id"]: n["row"] for n in normed}

    for cid in in_pair:
        root = find(cid)
        by_root.setdefault(root, set()).add(cid)

    groups = []
    for root, ids in by_root.items():
        members = []
        for cid in ids:
            row = id_to_row.get(cid, {})
            # Find best pairwise score this customer has in the group
            best = 0
            best_type = ""
            for other_id, meta in pair_index.get(cid, []):
                if other_id in ids and meta["score"] > best:
                    best = meta["score"]
                    best_type = meta["match_type"]
            members.append({
                "id": cid,
                "company_name": row.get("company_name") or "",
                "website": row.get("website") or "",
                "city": row.get("city") or "",
                "district": row.get("district") or "",
                "market": row.get("market") or "",
                "application": row.get("application") or "",
                "status": row.get("status") or "",
                "partner": row.get("partner") or "",
                "competitor": row.get("competitor") or "",
                "phone": (row.get("contact_info") or {}).get("phone", "") if isinstance(row.get("contact_info"), dict) else "",
                "email": (row.get("contact_info") or {}).get("email", "") if isinstance(row.get("contact_info"), dict) else "",
                "created_at": row.get("created_at") or "",
                "best_score": best,
                "best_match_type": best_type,
            })
        # Sort members: most-filled first (these become natural "keep" suggestions)
        def _fillcount(m):
            return sum(1 for v in m.values() if v not in ("", None, [], {}))
        members.sort(key=_fillcount, reverse=True)
        max_score = max((m["best_score"] for m in members), default=0)
        groups.append({
            "id": root,
            "members": members,
            "size": len(members),
            "max_score": max_score,
        })

    # Sort: highest confidence first
    groups.sort(key=lambda g: -g["max_score"])

    result = {
        "threshold": threshold,
        "total_groups": len(groups),
        "total_customers_affected": sum(g["size"] for g in groups),
        "groups": groups,
    }

    _duplicates_cache["data"] = result
    _duplicates_cache["ts"] = now_ts
    _duplicates_cache["threshold"] = threshold
    return result


class DedupeMergeRequest(BaseModel):
    keep_id: str          # ID of the customer to keep
    delete_ids: List[str] # IDs to be merged & deleted
    field_overrides: Optional[Dict[str, Any]] = None  # manual choices {field: value}


@api_router.post("/customers-merge")
async def dedupe_merge(payload: DedupeMergeRequest, request: Request,
                       session_token: Optional[str] = Cookie(None)):
    """Merge duplicate customers into `keep_id` and delete the rest.

    - Visits, calls, activity_log entries are reassigned to keep_id.
    - Empty fields on keep are auto-filled from the most complete delete record.
    - `field_overrides` lets the UI force specific values per field.
    """
    user = await get_current_user_from_request(request, session_token)
    if not check_admin_permission(user):
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")

    if not payload.keep_id or not payload.delete_ids:
        raise HTTPException(status_code=400, detail="keep_id ve delete_ids zorunludur")
    if payload.keep_id in payload.delete_ids:
        raise HTTPException(status_code=400, detail="keep_id silinecek listede olamaz")

    # Fetch all involved customers
    all_ids = [payload.keep_id] + payload.delete_ids
    resp = supabase.table("customers").select("*").in_("id", all_ids).execute()
    customers = {c["id"]: c for c in (resp.data or [])}
    if payload.keep_id not in customers:
        raise HTTPException(status_code=404, detail="Hedef müşteri bulunamadı")
    missing = [d for d in payload.delete_ids if d not in customers]
    if missing:
        raise HTTPException(status_code=404, detail=f"Silinecek müşteri bulunamadı: {missing}")

    keep = customers[payload.keep_id]
    sources = [customers[d] for d in payload.delete_ids]

    # Auto-fill: take a value from the first source where keep is empty
    def _empty(v):
        if v is None: return True
        if isinstance(v, str): return not v.strip()
        if isinstance(v, (list, dict)): return len(v) == 0
        return False

    mergeable_fields = [
        "market", "application", "city", "district", "website",
        "competitor", "partner", "potential_level", "potential_value",
        "assigned_to", "notes", "description", "status",
        "next_followup_date", "is_followup",
    ]

    update_data = {}
    for field in mergeable_fields:
        if not _empty(keep.get(field)):
            continue
        for src in sources:
            if not _empty(src.get(field)):
                update_data[field] = src.get(field)
                break

    # contact_info
    keep_contact = keep.get("contact_info") or {}
    merged_contact = dict(keep_contact)
    for key in ("contact_person", "email", "phone"):
        if not merged_contact.get(key):
            for src in sources:
                src_contact = src.get("contact_info") or {}
                if src_contact.get(key):
                    merged_contact[key] = src_contact[key]
                    break
    if merged_contact != keep_contact:
        update_data["contact_info"] = merged_contact

    # contacts (list of additional contacts) — union
    keep_contacts = keep.get("contacts") or []
    src_contacts_union = list(keep_contacts)
    for src in sources:
        for c in (src.get("contacts") or []):
            if c not in src_contacts_union:
                src_contacts_union.append(c)
    if src_contacts_union != keep_contacts:
        update_data["contacts"] = src_contacts_union

    # products & tags — union of arrays
    for arr_field in ("products", "tags"):
        merged_set = list(keep.get(arr_field) or [])
        for src in sources:
            for v in (src.get(arr_field) or []):
                if v not in merged_set:
                    merged_set.append(v)
        if merged_set != (keep.get(arr_field) or []):
            update_data[arr_field] = merged_set

    # documents & notes_list — concat
    for arr_field in ("documents", "notes_list"):
        combined = list(keep.get(arr_field) or [])
        for src in sources:
            combined.extend(src.get(arr_field) or [])
        if combined != (keep.get(arr_field) or []):
            update_data[arr_field] = combined

    # Manual overrides — UI can force a specific value per field
    if payload.field_overrides:
        for k, v in payload.field_overrides.items():
            update_data[k] = v

    # Update keep
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        supabase.table("customers").update(update_data).eq("id", payload.keep_id).execute()

    # Reassign related records
    reassign_counts = {"visits": 0, "calls": 0, "activity_log": 0}
    for d_id in payload.delete_ids:
        for table, key in (("visits", "customer_id"), ("calls", "customer_id"), ("activity_log", "customer_id")):
            try:
                # update returns the affected rows
                r = supabase.table(table).update({key: payload.keep_id}).eq(key, d_id).execute()
                if r and r.data:
                    reassign_counts[table] += len(r.data)
            except Exception as e:
                logging.warning(f"Reassign {table} failed for {d_id}: {e}")

    # Delete source customers
    for d_id in payload.delete_ids:
        supabase.table("customers").delete().eq("id", d_id).execute()

    # Invalidate caches
    _invalidate_kanban_cache()
    _stats_cache["data"] = None
    _filter_options_cache["data"] = None
    _duplicates_cache["data"] = None

    # Log activity
    deleted_names = ", ".join(customers[d].get("company_name", "?") for d in payload.delete_ids)
    await log_activity(
        activity_type="customer_updated",
        title=f"Yinelenenler birleştirildi: {keep.get('company_name')}",
        subtitle=f"Silinen: {deleted_names}  ·  {sum(reassign_counts.values())} ilişkili kayıt taşındı",
        customer_id=payload.keep_id,
        customer_name=keep.get("company_name"),
    )

    return {
        "ok": True,
        "keep_id": payload.keep_id,
        "deleted_ids": payload.delete_ids,
        "updated_fields": list(update_data.keys()),
        "reassigned": reassign_counts,
    }


class AutoMergeRequest(BaseModel):
    min_score: float = 98.0
    dry_run: bool = False
    max_groups: int = 500


@api_router.post("/customers-auto-merge")
async def auto_merge_duplicates(payload: AutoMergeRequest, request: Request,
                                session_token: Optional[str] = Cookie(None)):
    """Bulk-merge all duplicate groups whose top score is >= min_score.

    For each qualifying group: keeps the most-filled customer, merges
    the rest into it (visits/calls/activity_log reassigned, empty fields
    auto-filled, arrays union'd), deletes the duplicates.

    Set `dry_run=true` to preview without making any changes.
    """
    user = await get_current_user_from_request(request, session_token)
    if not check_admin_permission(user):
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")

    # IMPORTANT: We must scan with a LOW threshold (e.g. 90) and then filter the
    # qualifying groups locally by max_score >= min_score.
    #
    # Reason: the pair-detection loop checks `name_sim >= thresh` BEFORE applying
    # context boosts (same city / phone / website). With threshold=98, two
    # records like "elzer otomasyon" vs "elzerotomasyon" (name_sim ~0.96) are
    # filtered out — even though their composite score WOULD be 100 after the
    # city/website boost. Scanning at a low threshold ensures we don't miss
    # these high-confidence pairs.
    SCAN_THRESHOLD = 90
    # Force fresh scan if cached value uses a different threshold.
    if _duplicates_cache.get("threshold") != SCAN_THRESHOLD:
        _duplicates_cache["data"] = None
    scan = await find_duplicate_customers(
        threshold=SCAN_THRESHOLD, request=request, session_token=session_token
    )
    qualifying = [g for g in scan["groups"] if g["max_score"] >= payload.min_score]
    qualifying = qualifying[: payload.max_groups]

    if payload.dry_run:
        return {
            "ok": True,
            "dry_run": True,
            "min_score": payload.min_score,
            "groups_eligible": len(qualifying),
            "customers_to_delete": sum(g["size"] - 1 for g in qualifying),
            "preview": [
                {
                    "keep": g["members"][0]["company_name"],
                    "delete": [m["company_name"] for m in g["members"][1:]],
                    "score": g["max_score"],
                }
                for g in qualifying[:20]
            ],
        }

    merged_groups = 0
    deleted_total = 0
    reassigned_total = 0
    errors = []

    for g in qualifying:
        if g["size"] < 2:
            continue
        # The endpoint already sorted members by fill count desc, so [0] is most filled.
        keep_id = g["members"][0]["id"]
        delete_ids = [m["id"] for m in g["members"][1:]]
        try:
            result = await dedupe_merge(
                DedupeMergeRequest(keep_id=keep_id, delete_ids=delete_ids),
                request=request,
                session_token=session_token,
            )
            merged_groups += 1
            deleted_total += len(delete_ids)
            reassigned = result.get("reassigned", {})
            reassigned_total += sum(reassigned.values())
        except Exception as e:
            errors.append({
                "keep": g["members"][0]["company_name"],
                "error": str(e),
            })

    return {
        "ok": True,
        "dry_run": False,
        "min_score": payload.min_score,
        "groups_merged": merged_groups,
        "customers_deleted": deleted_total,
        "records_reassigned": reassigned_total,
        "errors": errors,
    }


class BulkDeleteRequest(BaseModel):
    ids: List[str]


class BulkUpdateRequest(BaseModel):
    ids: List[str]
    updates: Dict[str, Any]


@api_router.post("/customers-bulk-delete")
async def bulk_delete_customers(payload: BulkDeleteRequest, request: Request,
                                session_token: Optional[str] = Cookie(None)):
    """Delete many customers in one go. Admin only."""
    user = await get_current_user_from_request(request, session_token)
    if not check_admin_permission(user):
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
    if not payload.ids:
        raise HTTPException(status_code=400, detail="ids zorunlu")
    if len(payload.ids) > 1000:
        raise HTTPException(status_code=400, detail="Tek seferde en fazla 1000 kayıt")

    # Supabase .in_ filter chunked for >100 ids (postgrest URL length safety)
    deleted = 0
    chunk = 100
    for i in range(0, len(payload.ids), chunk):
        batch = payload.ids[i:i+chunk]
        # Reassign related records to NULL or delete them? We delete the customer rows;
        # FK constraints in Supabase typically cascade or set null depending on schema.
        # Visits/calls/activity_log rows tied to deleted customers will be orphaned;
        # delete them explicitly to keep DB clean.
        for table in ("visits", "calls", "activity_log"):
            try:
                supabase.table(table).delete().in_("customer_id", batch).execute()
            except Exception as e:
                logging.warning(f"Bulk delete cleanup {table}: {e}")
        try:
            r = supabase.table("customers").delete().in_("id", batch).execute()
            deleted += len(r.data or [])
        except Exception as e:
            logging.error(f"Bulk delete customers chunk failed: {e}")

    _invalidate_kanban_cache()
    _stats_cache["data"] = None
    _filter_options_cache["data"] = None
    _latest_calls_cache["data"] = None
    return {"ok": True, "deleted": deleted}


@api_router.post("/customers-bulk-update")
async def bulk_update_customers(payload: BulkUpdateRequest, request: Request,
                                session_token: Optional[str] = Cookie(None)):
    """Update one or more fields across many customers. Admin only."""
    user = await get_current_user_from_request(request, session_token)
    if not check_admin_permission(user):
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
    if not payload.ids or not payload.updates:
        raise HTTPException(status_code=400, detail="ids ve updates zorunlu")
    if len(payload.ids) > 1000:
        raise HTTPException(status_code=400, detail="Tek seferde en fazla 1000 kayıt")

    # Whitelist updatable fields to prevent accidental id/created_at overwrite
    allowed = {
        "market", "application", "city", "district", "status", "is_followup",
        "next_followup_date", "partner", "competitor", "assigned_to",
        "potential_level", "potential_value", "notes", "description", "tags",
    }
    safe_updates = {k: v for k, v in payload.updates.items() if k in allowed}
    if not safe_updates:
        raise HTTPException(status_code=400, detail="Güncellenebilir alan yok")
    safe_updates["updated_at"] = datetime.now(timezone.utc).isoformat()

    updated = 0
    chunk = 100
    for i in range(0, len(payload.ids), chunk):
        batch = payload.ids[i:i+chunk]
        try:
            r = supabase.table("customers").update(safe_updates).in_("id", batch).execute()
            updated += len(r.data or [])
        except Exception as e:
            logging.error(f"Bulk update chunk failed: {e}")

    _invalidate_kanban_cache()
    _stats_cache["data"] = None
    _filter_options_cache["data"] = None
    return {"ok": True, "updated": updated, "fields": list(safe_updates.keys())}


class BulkExportRequest(BaseModel):
    ids: List[str]


@api_router.post("/customers-bulk-export")
async def bulk_export_customers(payload: BulkExportRequest, request: Request,
                                session_token: Optional[str] = Cookie(None)):
    """Export selected customers as XLSX. Admin only."""
    user = await get_current_user_from_request(request, session_token)
    if not check_admin_permission(user):
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
    if not payload.ids:
        raise HTTPException(status_code=400, detail="ids zorunlu")

    # Fetch in chunks to respect Supabase 1000-row cap
    rows = []
    chunk = 200
    for i in range(0, len(payload.ids), chunk):
        batch = payload.ids[i:i+chunk]
        r = supabase.table("customers").select("*").in_("id", batch).execute()
        rows.extend(r.data or [])

    # Build Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Müşteriler"
    headers = [
        "Firma Adı", "Market", "Uygulama", "Şehir", "İlçe", "Durum",
        "Partner", "Rakip", "Takip Eden", "Web Sitesi", "Telefon",
        "E-posta", "İlgili Kişi", "Potansiyel", "Notlar",
        "Eklenme Tarihi", "Son Güncelleme",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for c in rows:
        contact = c.get("contact_info") or {}
        ws.append([
            c.get("company_name", ""),
            c.get("market", ""),
            c.get("application", ""),
            c.get("city", ""),
            c.get("district", ""),
            c.get("status", ""),
            c.get("partner", ""),
            c.get("competitor", ""),
            c.get("assigned_to", ""),
            c.get("website", ""),
            contact.get("phone", "") if isinstance(contact, dict) else "",
            contact.get("email", "") if isinstance(contact, dict) else "",
            contact.get("contact_person", "") if isinstance(contact, dict) else "",
            c.get("potential_level", "") or c.get("potential_value", ""),
            c.get("notes", "") or c.get("description", ""),
            c.get("created_at", ""),
            c.get("updated_at", ""),
        ])
    for col in ws.columns:
        try:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"musteriler_secili_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ============ BULK IMPORT ENDPOINTS ============

@api_router.post("/import/bulk-customers")
async def bulk_import_customers(data: dict):
    """Bulk import customers - accepts a batch of customers and inserts them efficiently"""
    items = data.get("items", [])
    merge_items = data.get("merge_items", [])  # [{customer_id, data}]
    
    if not items and not merge_items:
        return {"success": 0, "failed": 0, "merged": 0}
    
    success_count = 0
    fail_count = 0
    merge_count = 0
    
    # Process new customers in batch
    if items:
        batch_docs = []
        for item in items:
            try:
                # Clean up import-specific fields
                item.pop("rowNum", None)
                item.pop("similarity_status", None)
                item.pop("similar_customers", None)
                
                # Ensure contact_info is a dict
                if item.get("contact_info") and hasattr(item["contact_info"], "items"):
                    item["contact_info"] = dict(item["contact_info"])
                elif not item.get("contact_info"):
                    item["contact_info"] = {"contact_person": "", "email": "", "phone": ""}
                
                # Set defaults
                doc = {
                    "id": str(uuid.uuid4()),
                    "company_name": item.get("company_name", ""),
                    "market": item.get("market", ""),
                    "application": item.get("application", ""),
                    "city": item.get("city", ""),
                    "district": item.get("district", ""),
                    "website": item.get("website", ""),
                    "status": item.get("status", "Beklemede"),
                    "contact_info": item.get("contact_info", {}),
                    "contacts": item.get("contacts", []),
                    "potential_value": item.get("potential_value", 0),
                    "next_followup_date": item.get("next_followup_date", ""),
                    "assigned_to": item.get("assigned_to", ""),
                    "competitor": item.get("competitor", ""),
                    "partner": item.get("partner", ""),
                    "potential_level": item.get("potential_level", "Düşük"),
                    "products": item.get("products", []),
                    "description": item.get("description", ""),
                    "notes": item.get("notes", ""),
                    "notes_list": item.get("notes_list", []),
                    "documents": item.get("documents", []),
                    "tags": item.get("tags", []),
                    "is_followup": item.get("is_followup", False),
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                batch_docs.append(doc)
            except Exception as e:
                logging.error(f"Error preparing customer: {e}")
                fail_count += 1
        
        # Insert in batches of 100 to avoid Supabase limits
        batch_size = 100
        for i in range(0, len(batch_docs), batch_size):
            batch = batch_docs[i:i + batch_size]
            try:
                supabase.table("customers").insert(batch).execute()
                success_count += len(batch)
            except Exception as e:
                logging.error(f"Batch insert error: {e}")
                # Fallback: insert one by one
                for doc in batch:
                    try:
                        supabase.table("customers").insert(doc).execute()
                        success_count += 1
                    except Exception as e2:
                        logging.error(f"Single insert error: {e2}")
                        fail_count += 1
    
    # Process merge items
    for merge_item in merge_items:
        try:
            customer_id = merge_item.get("customer_id")
            new_data = merge_item.get("data", {})
            
            # Clean up import-specific fields
            new_data.pop("rowNum", None)
            new_data.pop("similarity_status", None)
            new_data.pop("similar_customers", None)
            
            response = supabase.table("customers").select("*").eq("id", customer_id).execute()
            if not response.data:
                fail_count += 1
                continue
            
            existing = response.data[0]
            update_data = {}
            
            merge_fields = [
                "market", "application", "city", "district", "website",
                "competitor", "partner", "potential_level", "assigned_to",
                "notes", "description"
            ]
            
            for field in merge_fields:
                existing_value = existing.get(field, "")
                new_value = new_data.get(field, "")
                if (not existing_value or existing_value == "") and new_value:
                    update_data[field] = new_value
            
            # Handle contact_info
            existing_contact = existing.get("contact_info", {}) or {}
            new_contact = new_data.get("contact_info", {}) or {}
            merged_contact = {**existing_contact}
            for key in ["contact_person", "email", "phone"]:
                if (not existing_contact.get(key)) and new_contact.get(key):
                    merged_contact[key] = new_contact[key]
            if merged_contact != existing_contact:
                update_data["contact_info"] = merged_contact
            
            # Merge products
            existing_products = existing.get("products", []) or []
            new_products = new_data.get("products", []) or []
            if new_products:
                merged_products = list(set(existing_products + new_products))
                if merged_products != existing_products:
                    update_data["products"] = merged_products
            
            if update_data:
                update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
                supabase.table("customers").update(update_data).eq("id", customer_id).execute()
            
            merge_count += 1
        except Exception as e:
            logging.error(f"Merge error: {e}")
            fail_count += 1
    
    # Log bulk import activity
    await log_activity(
        activity_type="bulk_import",
        title=f"Toplu içe aktarma: {success_count} yeni, {merge_count} güncellendi",
        subtitle=f"{fail_count} başarısız" if fail_count > 0 else "",
        customer_id=None,
        customer_name=""
    )
    
    return {
        "success": success_count,
        "merged": merge_count,
        "failed": fail_count
    }

@api_router.post("/options/bulk")
async def bulk_create_options(data: dict):
    """Bulk create options - accepts a list of options and creates unique ones"""
    options = data.get("options", [])
    
    if not options:
        return {"created": 0, "skipped": 0}
    
    # Get all existing options first
    existing_response = supabase.table("options").select("field_name, value").execute()
    existing_set = set()
    for opt in existing_response.data:
        existing_set.add((opt["field_name"], opt["value"].lower()))
    
    # Filter new unique options
    new_options = []
    seen = set()
    for opt in options:
        field_name = opt.get("field_name", "")
        value = opt.get("value", "").strip()
        if not field_name or not value:
            continue
        
        key = (field_name, value.lower())
        if key not in existing_set and key not in seen:
            seen.add(key)
            new_options.append({
                "id": str(uuid.uuid4()),
                "field_name": field_name,
                "value": value,
                "color": opt.get("color")
            })
    
    created = 0
    if new_options:
        # Insert in batches
        batch_size = 100
        for i in range(0, len(new_options), batch_size):
            batch = new_options[i:i + batch_size]
            try:
                supabase.table("options").insert(batch).execute()
                created += len(batch)
            except Exception as e:
                logging.error(f"Bulk options insert error: {e}")
                # Fallback: insert one by one
                for opt in batch:
                    try:
                        supabase.table("options").insert(opt).execute()
                        created += 1
                    except Exception:
                        pass
    
    return {"created": created, "skipped": len(options) - created}


# ============ VISITS ENDPOINTS ============

@api_router.post("/visits", response_model=Visit)
async def create_visit(visit: VisitCreate):
    visit_obj = Visit(**visit.model_dump())
    doc = visit_obj.model_dump()
    
    supabase.table("visits").insert(doc).execute()
    
    # Get customer name
    customer_name = ""
    if visit.customer_id:
        customer = supabase.table("customers").select("company_name").eq("id", visit.customer_id).execute()
        if customer.data:
            customer_name = customer.data[0].get("company_name", "")
    
    # Log activity
    await log_activity(
        activity_type="visit_created",
        title=f"Ziyaret: {customer_name}",
        subtitle=visit.visit_type or "Ziyaret",
        customer_id=visit.customer_id,
        customer_name=customer_name
    )
    
    return visit_obj

@api_router.get("/visits", response_model=List[Visit])
async def get_visits(customer_id: Optional[str] = None):
    query = supabase.table("visits").select("*")
    if customer_id:
        query = query.eq("customer_id", customer_id)
    
    response = query.order("created_at", desc=True).execute()
    return response.data

@api_router.get("/visits/{visit_id}", response_model=Visit)
async def get_visit(visit_id: str):
    response = supabase.table("visits").select("*").eq("id", visit_id).execute()
    if not response.data:
        raise HTTPException(status_code=404, detail="Ziyaret bulunamadı")
    return response.data[0]

@api_router.put("/visits/{visit_id}", response_model=Visit)
async def update_visit(visit_id: str, visit: VisitCreate):
    response = supabase.table("visits").select("*").eq("id", visit_id).execute()
    if not response.data:
        raise HTTPException(status_code=404, detail="Ziyaret bulunamadı")
    
    update_data = visit.model_dump()
    supabase.table("visits").update(update_data).eq("id", visit_id).execute()
    
    updated = supabase.table("visits").select("*").eq("id", visit_id).execute()
    return updated.data[0]

@api_router.delete("/visits/{visit_id}")
async def delete_visit(visit_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user_from_request(request, session_token)
    if not check_permission(user, "can_delete"):
        raise HTTPException(status_code=403, detail="Bu işlem için silme yetkisi gerekli")
    
    supabase.table("visits").delete().eq("id", visit_id).execute()
    return {"message": "Ziyaret silindi"}

# ============ CALLS ENDPOINTS ============

@api_router.post("/calls", response_model=Call)
async def create_call(call: CallCreate):
    call_obj = Call(**call.model_dump())
    doc = call_obj.model_dump()
    
    supabase.table("calls").insert(doc).execute()
    _latest_calls_cache["data"] = None  # invalidate

    # Get customer name
    customer_name = ""
    if call.customer_id:
        customer = supabase.table("customers").select("company_name").eq("id", call.customer_id).execute()
        if customer.data:
            customer_name = customer.data[0].get("company_name", "")
    
    # Log activity
    await log_activity(
        activity_type="call_created",
        title=f"Arama: {customer_name}",
        subtitle=f"{call.caller_name or ''} - {call.outcome or ''}",
        customer_id=call.customer_id,
        customer_name=customer_name
    )
    
    return call_obj

@api_router.get("/calls", response_model=List[Call])
async def get_calls(customer_id: Optional[str] = None):
    query = supabase.table("calls").select("*")
    if customer_id:
        query = query.eq("customer_id", customer_id)
    
    response = query.order("created_at", desc=True).execute()
    return response.data

@api_router.get("/calls/{call_id}", response_model=Call)
async def get_call(call_id: str):
    response = supabase.table("calls").select("*").eq("id", call_id).execute()
    if not response.data:
        raise HTTPException(status_code=404, detail="Arama bulunamadı")
    return response.data[0]

@api_router.put("/calls/{call_id}", response_model=Call)
async def update_call(call_id: str, call: CallCreate):
    response = supabase.table("calls").select("*").eq("id", call_id).execute()
    if not response.data:
        raise HTTPException(status_code=404, detail="Arama bulunamadı")
    
    update_data = call.model_dump()
    supabase.table("calls").update(update_data).eq("id", call_id).execute()
    _latest_calls_cache["data"] = None

    updated = supabase.table("calls").select("*").eq("id", call_id).execute()
    return updated.data[0]

@api_router.delete("/calls/{call_id}")
async def delete_call(call_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user_from_request(request, session_token)
    if not check_permission(user, "can_delete"):
        raise HTTPException(status_code=403, detail="Bu işlem için silme yetkisi gerekli")
    
    supabase.table("calls").delete().eq("id", call_id).execute()
    _latest_calls_cache["data"] = None
    return {"message": "Arama silindi"}

@api_router.get("/customers/{customer_id}/calls", response_model=List[Call])
async def get_customer_calls(customer_id: str):
    response = supabase.table("calls").select("*").eq("customer_id", customer_id).order("created_at", desc=True).execute()
    return response.data

# ============ OPTIONS ENDPOINTS ============

@api_router.get("/options/grouped")
async def get_grouped_options():
    # Fetch all options with pagination to handle >1000 records
    all_options = []
    offset = 0
    limit = 1000
    while True:
        response = supabase.table("options").select("*").range(offset, offset + limit - 1).execute()
        if not response.data:
            break
        all_options.extend(response.data)
        if len(response.data) < limit:
            break
        offset += limit
    
    grouped = {}
    for opt in all_options:
        field = opt["field_name"]
        if field not in grouped:
            grouped[field] = []
        grouped[field].append(opt)
    return grouped

@api_router.get("/options/{field_name}", response_model=List[DynamicOption])
async def get_options(field_name: str):
    # Fetch all options for field with pagination to handle >1000 records
    all_options = []
    offset = 0
    limit = 1000
    while True:
        response = supabase.table("options").select("*").eq("field_name", field_name).range(offset, offset + limit - 1).execute()
        if not response.data:
            break
        all_options.extend(response.data)
        if len(response.data) < limit:
            break
        offset += limit
    return all_options

@api_router.get("/options", response_model=List[DynamicOption])
async def get_all_options():
    # Fetch all options with pagination to handle >1000 records
    all_options = []
    offset = 0
    limit = 1000
    while True:
        response = supabase.table("options").select("*").range(offset, offset + limit - 1).execute()
        if not response.data:
            break
        all_options.extend(response.data)
        if len(response.data) < limit:
            break
        offset += limit
    return all_options

@api_router.post("/options", response_model=DynamicOption)
async def create_option(option: DynamicOption):
    # Check if exists (case-insensitive) - handle pagination for >1000 records
    all_field_options = []
    offset = 0
    limit = 1000
    while True:
        response = supabase.table("options").select("*").eq("field_name", option.field_name).range(offset, offset + limit - 1).execute()
        if not response.data:
            break
        all_field_options.extend(response.data)
        if len(response.data) < limit:
            break
        offset += limit
    
    existing = [o for o in all_field_options if o.get("value", "").lower() == option.value.lower()]
    if existing:
        return existing[0]
    
    doc = option.model_dump()
    supabase.table("options").insert(doc).execute()
    return option

@api_router.put("/options/{option_id}")
async def update_option(option_id: str, option_data: dict):
    response = supabase.table("options").select("*").eq("id", option_id).execute()
    if not response.data:
        raise HTTPException(status_code=404, detail="Seçenek bulunamadı")
    
    # Only update value and color
    update_data = {}
    if "value" in option_data:
        update_data["value"] = option_data["value"]
    if "color" in option_data:
        update_data["color"] = option_data["color"]
    
    if update_data:
        supabase.table("options").update(update_data).eq("id", option_id).execute()
    
    updated = supabase.table("options").select("*").eq("id", option_id).execute()
    return updated.data[0]

@api_router.delete("/options/{option_id}")
async def delete_option(option_id: str):
    # Allow all authenticated users to delete options (no admin check needed for dropdown options)
    response = supabase.table("options").select("*").eq("id", option_id).execute()
    if not response.data:
        raise HTTPException(status_code=404, detail="Seçenek bulunamadı")
    
    supabase.table("options").delete().eq("id", option_id).execute()
    return {"message": "Seçenek silindi"}

# ============ SAVED FILTERS ENDPOINTS ============

@api_router.get("/filters", response_model=List[SavedFilter])
async def get_saved_filters():
    response = supabase.table("saved_filters").select("*").execute()
    return response.data

@api_router.post("/filters", response_model=SavedFilter)
async def create_saved_filter(filter_data: SavedFilterCreate):
    filter_obj = SavedFilter(**filter_data.model_dump())
    doc = filter_obj.model_dump()
    doc["conditions"] = [c.model_dump() if hasattr(c, 'model_dump') else c for c in doc["conditions"]]
    
    supabase.table("saved_filters").insert(doc).execute()
    return filter_obj

@api_router.put("/filters/{filter_id}", response_model=SavedFilter)
async def update_saved_filter(filter_id: str, filter_data: SavedFilterCreate):
    response = supabase.table("saved_filters").select("*").eq("id", filter_id).execute()
    if not response.data:
        raise HTTPException(status_code=404, detail="Filtre bulunamadı")
    
    update_data = filter_data.model_dump()
    update_data["conditions"] = [c.model_dump() if hasattr(c, 'model_dump') else c for c in update_data["conditions"]]
    
    supabase.table("saved_filters").update(update_data).eq("id", filter_id).execute()
    
    updated = supabase.table("saved_filters").select("*").eq("id", filter_id).execute()
    return updated.data[0]

@api_router.delete("/filters/{filter_id}")
async def delete_saved_filter(filter_id: str):
    supabase.table("saved_filters").delete().eq("id", filter_id).execute()
    return {"message": "Filtre silindi"}

@api_router.post("/filters/{filter_id}/apply")
async def apply_saved_filter(filter_id: str):
    response = supabase.table("saved_filters").select("*").eq("id", filter_id).execute()
    if not response.data:
        raise HTTPException(status_code=404, detail="Filtre bulunamadı")
    
    filter_doc = response.data[0]
    conditions = filter_doc.get("conditions", [])
    logic = filter_doc.get("logic", "AND")
    
    # Get all customers first
    customers_response = supabase.table("customers").select("*").execute()
    customers = customers_response.data
    
    if not conditions:
        return customers
    
    # Apply filters in Python for complex conditions
    def matches_condition(customer, cond):
        field = cond.get("field")
        operator = cond.get("operator")
        value = cond.get("value", "")
        
        customer_value = customer.get(field, "")
        if customer_value is None:
            customer_value = ""
        
        if operator == "equals":
            return str(customer_value).lower() == str(value).lower()
        elif operator == "contains":
            return str(value).lower() in str(customer_value).lower()
        elif operator == "not_equals":
            return str(customer_value).lower() != str(value).lower()
        elif operator == "greater_than":
            try:
                return float(customer_value or 0) > float(value)
            except (TypeError, ValueError):
                return False
        elif operator == "less_than":
            try:
                return float(customer_value or 0) < float(value)
            except (TypeError, ValueError):
                return False
        elif operator == "is_empty":
            return not customer_value
        elif operator == "is_not_empty":
            return bool(customer_value)
        return False
    
    filtered = []
    for customer in customers:
        if logic == "OR":
            if any(matches_condition(customer, c) for c in conditions):
                filtered.append(customer)
        else:  # AND
            if all(matches_condition(customer, c) for c in conditions):
                filtered.append(customer)
    
    return filtered

# ============ FILE UPLOAD ENDPOINTS ============

@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...), customer_id: str = None):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Dosya adı gerekli")
    
    file_ext = Path(file.filename).suffix
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = UPLOADS_DIR / unique_filename
    
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Dosya kaydedilemedi: {str(e)}")
    
    file_info = {
        "id": str(uuid.uuid4()),
        "original_name": file.filename,
        "stored_name": unique_filename,
        "url": f"/api/files/{unique_filename}",
        "size": file_path.stat().st_size,
        "content_type": file.content_type,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    customer_name = ""
    if customer_id:
        response = supabase.table("customers").select("company_name, documents").eq("id", customer_id).execute()
        if response.data:
            customer_name = response.data[0].get("company_name", "")
            documents = response.data[0].get("documents") or []
            doc = {
                "id": file_info["id"],
                "name": file.filename,
                "url": file_info["url"],
                "stored_name": unique_filename,
                "size": file_info["size"],
                "created_at": file_info["created_at"]
            }
            documents.append(doc)
            supabase.table("customers").update({"documents": documents}).eq("id", customer_id).execute()
            
            # Log activity
            await log_activity(
                activity_type="file_uploaded",
                title=f"Dosya yüklendi: {file.filename}",
                subtitle=customer_name,
                customer_id=customer_id,
                customer_name=customer_name
            )
    
    return file_info

@api_router.get("/files/{filename}")
async def get_file(filename: str):
    file_path = UPLOADS_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Dosya bulunamadı")
    
    return FileResponse(file_path)

@api_router.delete("/files/{filename}")
async def delete_file(filename: str, customer_id: str = None):
    file_path = UPLOADS_DIR / filename
    
    if file_path.exists():
        file_path.unlink()
    
    if customer_id:
        response = supabase.table("customers").select("documents").eq("id", customer_id).execute()
        if response.data:
            documents = response.data[0].get("documents") or []
            documents = [d for d in documents if d.get("stored_name") != filename]
            supabase.table("customers").update({"documents": documents}).eq("id", customer_id).execute()
    
    return {"message": "Dosya silindi"}

# ============ FOLLOW-UP ENDPOINTS ============

@api_router.get("/followups")
async def get_followups():
    now = _time.time()
    if _followups_cache["data"] and (now - _followups_cache["ts"]) < _FOLLOWUPS_CACHE_TTL:
        return _followups_cache["data"]
    # Use lighter select for customers (no notes/description etc — bell only needs id/name/dates)
    customers_response = supabase.table("customers").select(
        "id, company_name, market, city, district, status, is_followup, "
        "next_followup_date, assigned_to, partner"
    ).eq("is_followup", True).execute()
    visits_response = supabase.table("visits").select("*").eq("is_followup", True).execute()

    visits = visits_response.data or []

    # Fetch all customer names in ONE query (was N+1)
    customer_ids = list({v.get("customer_id") for v in visits if v.get("customer_id")})
    name_map = {}
    if customer_ids:
        names_resp = supabase.table("customers").select("id, company_name").in_("id", customer_ids).execute()
        name_map = {c["id"]: c.get("company_name", "Bilinmiyor") for c in (names_resp.data or [])}

    for visit in visits:
        visit["customer_name"] = name_map.get(visit.get("customer_id"), "Bilinmiyor")

    result = {
        "customers": customers_response.data or [],
        "visits": visits
    }
    _followups_cache["data"] = result
    _followups_cache["ts"] = now
    return result

@api_router.patch("/customers/{customer_id}/followup")
async def update_customer_followup(customer_id: str, is_followup: bool = Query(...)):
    """Update customer follow-up status"""
    supabase.table("customers").update({
        "is_followup": is_followup,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }).eq("id", customer_id).execute()
    _followups_cache["data"] = None
    _stats_cache["data"] = None
    _invalidate_kanban_cache()
    return {"message": "Follow-up durumu güncellendi", "is_followup": is_followup}

@api_router.patch("/visits/{visit_id}/followup")
async def update_visit_followup(visit_id: str, is_followup: bool = Query(...)):
    """Update visit follow-up status"""
    supabase.table("visits").update({
        "is_followup": is_followup
    }).eq("id", visit_id).execute()
    _followups_cache["data"] = None
    return {"message": "Follow-up durumu güncellendi", "is_followup": is_followup}

# ============ DASHBOARD STATS ============

_stats_cache = {"data": None, "timestamp": 0}
_STATS_CACHE_TTL = 30  # 30 seconds cache

@api_router.get("/stats")
async def get_stats():
    now = _time.time()
    if _stats_cache["data"] and (now - _stats_cache["timestamp"]) < _STATS_CACHE_TTL:
        return _stats_cache["data"]
    
    # Only fetch needed fields - paginate to bypass Supabase 1000 row limit
    customers = fetch_all_rows(
        "customers",
        "id, status, market, city, is_followup, next_followup_date, company_name, created_at, assigned_to, partner",
    )
    visits_count_response = supabase.table("visits").select("id, is_followup", count="exact").execute()

    total_customers = len(customers)
    total_visits = visits_count_response.count or len(visits_count_response.data)
    followup_customers = sum(1 for c in customers if c.get("is_followup"))
    followup_visits = sum(1 for v in visits_count_response.data if v.get("is_followup"))
    
    # Status distribution
    status_dist = {}
    for c in customers:
        status = normalize_status(c.get("status", "")) or "Beklemede"
        status_dist[status] = status_dist.get(status, 0) + 1
    status_distribution = [{"_id": k, "count": v} for k, v in status_dist.items()]
    
    # Market distribution (top 10)
    market_dist = {}
    for c in customers:
        market = c.get("market", "")
        if market:
            market_dist[market] = market_dist.get(market, 0) + 1
    market_distribution = sorted([{"_id": k, "count": v} for k, v in market_dist.items()], key=lambda x: -x["count"])[:10]
    
    # City distribution (top 10)
    city_dist = {}
    for c in customers:
        city = c.get("city", "")
        if city:
            city_dist[city] = city_dist.get(city, 0) + 1
    city_distribution = sorted([{"_id": k, "count": v} for k, v in city_dist.items()], key=lambda x: -x["count"])[:10]

    # Assigned-to distribution (all customers) — top 10
    assigned_dist = {}
    for c in customers:
        a = (c.get("assigned_to") or "").strip()
        if a:
            assigned_dist[a] = assigned_dist.get(a, 0) + 1
    assigned_to_distribution = sorted(
        [{"_id": k, "count": v} for k, v in assigned_dist.items()],
        key=lambda x: -x["count"],
    )[:10]

    # Partner distribution among ONLY followup customers
    followup_partner_dist = {}
    for c in customers:
        if not c.get("is_followup"):
            continue
        p = (c.get("partner") or "").strip()
        if p:
            followup_partner_dist[p] = followup_partner_dist.get(p, 0) + 1
    partner_followup_distribution = sorted(
        [{"_id": k, "count": v} for k, v in followup_partner_dist.items()],
        key=lambda x: -x["count"],
    )[:10]
    
    # Recent customers (last 5) - already sorted by created_at desc from customers data
    recent_customers = sorted(customers, key=lambda x: x.get("created_at", ""), reverse=True)[:5]
    
    # Upcoming followups
    upcoming_followups = [c for c in customers if c.get("next_followup_date")]
    upcoming_followups = sorted(upcoming_followups, key=lambda x: x.get("next_followup_date", ""))[:5]
    
    result = {
        "total_customers": total_customers,
        "total_visits": total_visits,
        "followup_customers": followup_customers,
        "followup_visits": followup_visits,
        "status_distribution": status_distribution,
        "market_distribution": market_distribution,
        "city_distribution": city_distribution,
        "assigned_to_distribution": assigned_to_distribution,
        "partner_followup_distribution": partner_followup_distribution,
        "recent_customers": recent_customers,
        "upcoming_followups": upcoming_followups
    }
    
    _stats_cache["data"] = result
    _stats_cache["timestamp"] = now
    return result


# Distribution cache (per field+followup_only key) — 30s TTL.

# Data score (completeness) cache — invalidated on any customer write.
_data_score_cache = {"map": None, "ts": 0}
_DATA_SCORE_CACHE_TTL = 60

# Followups endpoint cache — 30s. Used by header bell on every page load.
_followups_cache = {"data": None, "ts": 0}
_FOLLOWUPS_CACHE_TTL = 30


@api_router.get("/stats/distribution")
async def get_stats_distribution(
    field: str = Query(..., description="Dimension to aggregate: market, status, city, assigned_to, partner, application, competitor"),
    followup_only: bool = Query(False),
    limit: int = Query(10, ge=1, le=50),
):
    """Aggregate count by field across all customers.

    Used by the dashboard's editable donut widgets. Lightweight: fetches only
    the needed columns, computes the distribution in-memory. Cached 30s per
    (field, followup_only, limit) tuple.
    """
    allowed = {"market", "status", "city", "assigned_to", "partner",
               "application", "competitor"}
    if field not in allowed:
        raise HTTPException(status_code=400, detail=f"field must be one of {sorted(allowed)}")

    now = _time.time()
    cache_key = f"{field}::{int(followup_only)}::{limit}"
    cached = _distribution_cache.get(cache_key)
    if cached and (now - cached["ts"]) < _DISTRIBUTION_CACHE_TTL:
        return cached["data"]

    # Select only the columns we need to keep this fast (paginated to bypass 1000 row limit)
    cols = f"id, is_followup, {field}"
    rows = fetch_all_rows("customers", cols)

    if followup_only:
        rows = [r for r in rows if r.get("is_followup")]

    dist: Dict[str, int] = {}
    for r in rows:
        raw_val = r.get(field)
        if field == "status":
            v = normalize_status(raw_val or "") or (raw_val or "")
        else:
            v = (raw_val or "").strip() if isinstance(raw_val, str) else raw_val
        if not v:
            continue
        dist[v] = dist.get(v, 0) + 1

    entries = sorted(
        [{"_id": k, "count": v} for k, v in dist.items()],
        key=lambda x: -x["count"],
    )[:limit]
    result = {
        "field": field,
        "followup_only": followup_only,
        "total": sum(e["count"] for e in entries),
        "entries": entries,
    }
    _distribution_cache[cache_key] = {"data": result, "ts": now}
    return result


@api_router.get("/stats/segment")
async def get_stats_segment(
    field: str = Query(..., description="Dimension to filter on (market, status, city, assigned_to, partner)"),
    value: str = Query(..., description="Value of that dimension"),
    followup_only: bool = Query(False, description="Restrict to is_followup=true customers"),
    limit: int = Query(100, ge=1, le=500),
):
    """Return the list of customers behind a specific chart segment.

    Used by the dashboard popup when the user clicks a slice in a donut chart.
    """
    allowed = {"market", "status", "city", "assigned_to", "partner", "application", "competitor"}
    if field not in allowed:
        raise HTTPException(status_code=400, detail=f"field must be one of {sorted(allowed)}")

    query = supabase.table("customers").select(
        "id, company_name, status, market, city, partner, assigned_to, is_followup"
    )

    if field == "status":
        # Status normalisation: also match the raw stored value
        normalized = normalize_status(value)
        if normalized != value:
            query = query.in_("status", [value, normalized])
        else:
            query = query.eq("status", value)
    else:
        query = query.eq(field, value)

    if followup_only:
        query = query.eq("is_followup", True)

    response = query.order("company_name").limit(limit).execute()
    return {
        "field": field,
        "value": value,
        "followup_only": followup_only,
        "count": len(response.data or []),
        "customers": response.data or [],
    }


@api_router.get("/activity-feed")
async def get_activity_feed(limit: int = 50):
    """Get recent activity feed from activity_log table (15s in-memory cache)."""
    now = _time.time()
    cache_key = f"activity_{limit}"
    cached = _activity_cache.get(cache_key)
    if cached and (now - cached["ts"]) < _ACTIVITY_CACHE_TTL:
        return cached["data"]
    try:
        response = supabase.table("activity_log").select("*").order("created_at", desc=True).limit(limit).execute()

        activities = []
        for log in response.data:
            activities.append({
                "id": log.get("id"),
                "type": log.get("activity_type", "unknown"),
                "title": log.get("title", ""),
                "subtitle": log.get("subtitle", ""),
                "timestamp": log.get("created_at", ""),
                "customer_id": log.get("customer_id"),
                "customer_name": log.get("customer_name", ""),
                "user_email": log.get("user_email", ""),
                "user_name": log.get("user_name", ""),
                "metadata": log.get("metadata", {})
            })

        _activity_cache[cache_key] = {"data": activities, "ts": now}
        return activities
    except Exception as e:
        logging.error(f"Activity feed error: {e}")
        # Fallback: return empty list if activity_log table doesn't exist yet
        return []


# ============ TEAM MEMBERS ENDPOINTS ============

@api_router.get("/team-members")
async def get_team_members():
    """
    List all distinct team members aggregated from:
      - customers.assigned_to
      - visits.visited_by
      - activity_log.user_name
    Returns one row per person with summary stats.
    """
    try:
        members = {}  # name -> stats

        def _bump(name, key, inc=1):
            n = (name or "").strip()
            if not n:
                return
            m = members.setdefault(n, {
                "name": n,
                "customers_count": 0,
                "followup_count": 0,
                "won_count": 0,
                "lost_count": 0,
                "visits_count": 0,
                "activities_count": 0,
                "last_activity": "",
            })
            m[key] = m.get(key, 0) + inc

        # Customers (paginated)
        cust_rows = fetch_all_rows("customers", "assigned_to, status, is_followup")
        for c in cust_rows:
            name = (c.get("assigned_to") or "").strip()
            if not name:
                continue
            _bump(name, "customers_count")
            if c.get("is_followup"):
                _bump(name, "followup_count")
            st = normalize_status(c.get("status", "") or "")
            if st == "Kazanıldı":
                _bump(name, "won_count")
            elif st == "Kaybedildi":
                _bump(name, "lost_count")

        # Visits (paginated)
        try:
            vis_rows = fetch_all_rows("visits", "visited_by, created_at")
            for v in vis_rows:
                name = (v.get("visited_by") or "").strip()
                if not name:
                    continue
                _bump(name, "visits_count")
                ts = v.get("created_at") or ""
                if ts and ts > members.get(name, {}).get("last_activity", ""):
                    members[name]["last_activity"] = ts
        except Exception:
            pass

        # Activity log (count + last_activity) — paginated
        try:
            act_rows = fetch_all_rows("activity_log", "user_name, created_at")
            for a in act_rows:
                name = (a.get("user_name") or "").strip()
                if not name:
                    continue
                _bump(name, "activities_count")
                ts = a.get("created_at") or ""
                if ts and ts > members.get(name, {}).get("last_activity", ""):
                    members[name]["last_activity"] = ts
        except Exception:
            pass

        result = sorted(members.values(), key=lambda m: -m["customers_count"])
        return {"members": result, "total": len(result)}
    except Exception as e:
        logging.error(f"team-members error: {e}")
        return {"members": [], "total": 0}


@api_router.get("/team-members/{name}/profile")
async def get_team_member_profile(name: str, days: int = 90, activity_limit: int = 100):
    """
    Detailed profile for a single team member.
    Returns:
      - summary stats
      - status distribution
      - market distribution
      - recent activities (from activity_log)
      - recent visits
      - assigned customers list
      - weekly activity (last N days)
    """
    try:
        name = (name or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="name required")

        # --- Customers assigned to this person (paginated) ---
        select_cols = (
            "id, company_name, status, market, city, is_followup, "
            "next_followup_date, partner, competitor, created_at, updated_at, assigned_to"
        )
        customers = []
        page_size = 1000
        offset = 0
        while True:
            cust_resp = supabase.table("customers").select(select_cols).ilike(
                "assigned_to", name
            ).range(offset, offset + page_size - 1).execute()
            chunk = cust_resp.data or []
            customers.extend(chunk)
            if len(chunk) < page_size:
                break
            offset += page_size

        # Stats
        won = sum(1 for c in customers if normalize_status(c.get("status", "")) == "Kazanıldı")
        lost = sum(1 for c in customers if normalize_status(c.get("status", "")) == "Kaybedildi")
        followups = sum(1 for c in customers if c.get("is_followup"))

        # Status distribution
        status_dist = {}
        for c in customers:
            st = normalize_status(c.get("status", "")) or "Beklemede"
            status_dist[st] = status_dist.get(st, 0) + 1
        status_distribution = sorted(
            [{"_id": k, "count": v} for k, v in status_dist.items()],
            key=lambda x: -x["count"],
        )

        # Market distribution
        market_dist = {}
        for c in customers:
            mk = (c.get("market") or "").strip()
            if mk:
                market_dist[mk] = market_dist.get(mk, 0) + 1
        market_distribution = sorted(
            [{"_id": k, "count": v} for k, v in market_dist.items()],
            key=lambda x: -x["count"],
        )[:10]

        # City distribution
        city_dist = {}
        for c in customers:
            ci = (c.get("city") or "").strip()
            if ci:
                city_dist[ci] = city_dist.get(ci, 0) + 1
        city_distribution = sorted(
            [{"_id": k, "count": v} for k, v in city_dist.items()],
            key=lambda x: -x["count"],
        )[:10]

        # --- Activities (from activity_log) ---
        activities = []
        try:
            act_resp = supabase.table("activity_log").select("*").ilike(
                "user_name", name
            ).order("created_at", desc=True).limit(activity_limit).execute()
            activities = act_resp.data or []
        except Exception:
            activities = []

        # --- Visits made by this person ---
        visits = []
        try:
            vis_resp = supabase.table("visits").select("*").ilike(
                "visited_by", name
            ).order("created_at", desc=True).limit(100).execute()
            visits = vis_resp.data or []
        except Exception:
            visits = []

        # --- Weekly activity buckets (last `days` days, daily counts) ---
        from collections import OrderedDict
        from datetime import timedelta
        today = datetime.now(timezone.utc).date()
        buckets = OrderedDict()
        for i in range(days - 1, -1, -1):
            d = today - timedelta(days=i)
            buckets[d.isoformat()] = 0

        for a in activities:
            ts = a.get("created_at") or ""
            if ts:
                try:
                    d = ts[:10]
                    if d in buckets:
                        buckets[d] += 1
                except Exception:
                    pass
        for v in visits:
            ts = v.get("created_at") or v.get("visit_date") or ""
            if ts:
                d = ts[:10]
                if d in buckets:
                    buckets[d] += 1

        activity_trend = [{"date": k, "count": v} for k, v in buckets.items()]

        # Last activity timestamp
        last_act = ""
        if activities:
            last_act = activities[0].get("created_at", "")
        for v in visits:
            ts = v.get("created_at", "")
            if ts and ts > last_act:
                last_act = ts

        return {
            "name": name,
            "summary": {
                "customers_count": len(customers),
                "followup_count": followups,
                "won_count": won,
                "lost_count": lost,
                "visits_count": len(visits),
                "activities_count": len(activities),
                "last_activity": last_act,
            },
            "status_distribution": status_distribution,
            "market_distribution": market_distribution,
            "city_distribution": city_distribution,
            "activity_trend": activity_trend,
            "activities": activities,
            "visits": visits,
            "customers": customers,
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"team-member profile error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# Lightweight health/warmup endpoint — used by external pingers (UptimeRobot,
# cron-job.org) to keep the Render free-tier container awake. Returns fast
# without hitting the DB, so it's safe to ping every few minutes.
@api_router.get("/health")
async def health():
    return {"status": "ok", "service": "crmaster-backend", "ts": int(_time.time())}

# ============ AUTH ENDPOINTS ============

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    user_id: str
    email: str
    name: str
    picture: Optional[str] = ""
    role: str = "user"
    auth_type: str = "google"
    password_hash: Optional[str] = None
    notification_days: int = 3
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

ADMIN_EMAIL = "hakanonel05@gmail.com"

def get_user_role(email: str) -> str:
    return "admin" if email == ADMIN_EMAIL else "user"

class SessionData(BaseModel):
    user_id: str
    email: str
    name: str
    picture: Optional[str] = ""
    role: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    expires_at: str = Field(default_factory=lambda: (datetime.now(timezone.utc) + timedelta(days=7)).isoformat())

sessions: Dict[str, SessionData] = {}

# Session persistence file - survives backend restarts.
# Render free tier filesystem is read-only outside the project tree and
# *ephemeral* within it. We attempt to write to the backend dir, but
# silently fall back to in-memory only if the filesystem rejects writes.
SESSIONS_FILE = Path(
    os.environ.get("SESSIONS_FILE") or (Path(__file__).parent / "sessions.json")
)

def _save_sessions_to_disk():
    """Save all sessions to disk for persistence across restarts"""
    try:
        data = {token: s.model_dump() for token, s in sessions.items()}
        SESSIONS_FILE.write_text(json.dumps(data))
    except (PermissionError, OSError) as e:
        # Read-only filesystem (e.g. some hosting envs) — silently degrade
        logging.debug(f"Sessions: skip disk save ({e})")
    except Exception as e:
        logging.warning(f"Failed to save sessions to disk: {e}")

def _load_sessions_from_disk():
    """Load sessions from disk on startup"""
    global sessions
    try:
        if SESSIONS_FILE.exists():
            data = json.loads(SESSIONS_FILE.read_text())
            now = datetime.now(timezone.utc)
            loaded = 0
            for token, sd in data.items():
                try:
                    session = SessionData(**sd)
                    # Skip expired sessions
                    if datetime.fromisoformat(session.expires_at.replace('Z', '+00:00')) > now:
                        sessions[token] = session
                        loaded += 1
                except Exception:
                    continue
            logging.info(f"Loaded {loaded} sessions from disk")
    except Exception as e:
        logging.warning(f"Failed to load sessions from disk: {e}")

def _persist_session(token: str, session: SessionData):
    """Persist sessions to disk"""
    sessions[token] = session
    _save_sessions_to_disk()

def _delete_persisted_session(token: str):
    """Remove session and persist"""
    if token in sessions:
        del sessions[token]
    _save_sessions_to_disk()

def _load_session_from_db(token: str) -> Optional[SessionData]:
    """Legacy alias - sessions now loaded from disk on startup"""
    return None

# Load existing sessions on module init
_load_sessions_from_disk()

def check_admin_permission(user: Optional[dict]) -> bool:
    if not user:
        return False
    user_email = (user.get("email") or "").lower()
    admin_email = ADMIN_EMAIL.lower()
    return user.get("role") == "admin" or user_email == admin_email


def check_super_admin(user: Optional[dict]) -> bool:
    """
    Super admin = ONLY the hardcoded ADMIN_EMAIL.
    Even users with role='admin' do NOT count as super admin.
    Used to gate user management & permission editing endpoints.
    """
    if not user:
        return False
    return (user.get("email") or "").lower() == ADMIN_EMAIL.lower()


# ============ USER PERMISSIONS ============
# Stored on disk because we cannot run DDL against the managed Supabase DB.
# Structure: { user_id: {"can_delete": bool, "can_edit_dashboard": bool} }
PERMISSIONS_FILE = Path(
    os.environ.get("PERMISSIONS_FILE") or (Path(__file__).parent / "user_permissions.json")
)
ALL_PERMISSION_KEYS = ("can_delete", "can_edit_dashboard")


def _load_permissions() -> dict:
    try:
        if PERMISSIONS_FILE.exists():
            with open(PERMISSIONS_FILE, "r") as f:
                return json.load(f) or {}
    except Exception as e:
        logging.warning(f"Could not load user permissions file: {e}")
    return {}


def _save_permissions(perms: dict) -> None:
    try:
        with open(PERMISSIONS_FILE, "w") as f:
            json.dump(perms, f)
    except Exception as e:
        logging.warning(f"Could not save user permissions file: {e}")


_user_permissions_cache: dict = _load_permissions()


def get_user_permissions(user_id: str) -> dict:
    p = _user_permissions_cache.get(user_id, {}) or {}
    return {k: bool(p.get(k, False)) for k in ALL_PERMISSION_KEYS}


def set_user_permission(user_id: str, key: str, value: bool) -> dict:
    if key not in ALL_PERMISSION_KEYS:
        raise ValueError(f"Unknown permission key: {key}")
    current = _user_permissions_cache.get(user_id, {}) or {}
    current[key] = bool(value)
    _user_permissions_cache[user_id] = current
    _save_permissions(_user_permissions_cache)
    return get_user_permissions(user_id)


def check_permission(user: Optional[dict], permission_key: str) -> bool:
    """Returns True if user is admin OR has the given permission flag set."""
    if not user:
        return False
    if check_admin_permission(user):
        return True
    uid = user.get("user_id") or user.get("id")
    if not uid:
        return False
    return bool(_user_permissions_cache.get(uid, {}).get(permission_key, False))


async def get_current_user_from_request(request: Request, session_token: Optional[str] = None) -> Optional[dict]:
    # Priority: explicit param > X-Session-Token header > cookie
    token = session_token or request.headers.get("x-session-token") or request.cookies.get("session_token")
    if not token:
        return None
    
    # Try in-memory first
    session = sessions.get(token)
    
    # Fallback to DB (for cases where backend restarted)
    if not session:
        session = _load_session_from_db(token)
        if session:
            sessions[token] = session  # Restore to in-memory cache
    
    if not session:
        return None
    
    try:
        if datetime.fromisoformat(session.expires_at.replace('Z', '+00:00')) < datetime.now(timezone.utc):
            if token in sessions:
                del sessions[token]
            _delete_persisted_session(token)
            return None
    except Exception:
        pass
    
    return {
        "user_id": session.user_id,
        "email": session.email,
        "name": session.name,
        "picture": session.picture,
        "role": session.role
    }

@api_router.get("/auth/me")
async def get_current_user(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user_from_request(request, session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Oturum açılmamış")
    # Attach effective permissions (admin → all true)
    if check_admin_permission(user):
        user["permissions"] = {k: True for k in ALL_PERMISSION_KEYS}
        user["is_admin"] = True
    else:
        user["permissions"] = get_user_permissions(user.get("user_id") or user.get("id"))
        user["is_admin"] = False
    user["is_super_admin"] = check_super_admin(user)
    return user

@api_router.post("/auth/callback")
async def auth_callback(request: Request, response: Response):
    data = await request.json()
    
    user_id = data.get("user_id", f"user_{uuid.uuid4().hex[:12]}")
    email = data.get("email", "")
    name = data.get("name", "")
    picture = data.get("picture", "")
    
    # Check whitelist - only approved emails can login
    whitelist_response = supabase.table("allowed_users").select("*").eq("email", email.lower()).execute()
    
    # Admin email always allowed
    admin_email = "hakanonel05@gmail.com"
    is_allowed = email.lower() == admin_email.lower() or (whitelist_response.data and len(whitelist_response.data) > 0)
    
    if not is_allowed:
        raise HTTPException(
            status_code=403, 
            detail="Bu e-posta adresi için erişim izni yok. Lütfen yöneticiyle iletişime geçin."
        )
    
    role = get_user_role(email)
    
    # Check if user exists in DB
    existing = supabase.table("users").select("*").eq("email", email).execute()
    
    if existing.data:
        user_doc = existing.data[0]
        supabase.table("users").update({
            "name": name,
            "picture": picture
        }).eq("email", email).execute()
    else:
        user_doc = {
            "user_id": user_id,
            "email": email,
            "name": name,
            "picture": picture,
            "role": role,
            "auth_type": "google",
            "notification_days": 3,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        supabase.table("users").insert(user_doc).execute()
    
    session_token = f"session_{uuid.uuid4().hex}"
    session_data = SessionData(
        user_id=user_id,
        email=email,
        name=name,
        picture=picture,
        role=role
    )
    sessions[session_token] = session_data
    _save_sessions_to_disk()
    
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        max_age=7*24*60*60
    )
    
    return {
        "user_id": user_id,
        "email": email,
        "name": name,
        "picture": picture,
        "role": role,
        "session_token": session_token
    }

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response, session_token: Optional[str] = Cookie(None)):
    if session_token and session_token in sessions:
        del sessions[session_token]
    if session_token:
        _delete_persisted_session(session_token)
    
    response.delete_cookie(key="session_token")
    return {"message": "Çıkış yapıldı"}

# ============ ALLOWED USERS (WHITELIST) ENDPOINTS ============

@api_router.get("/allowed-users")
async def get_allowed_users():
    """Get list of allowed users (whitelist)"""
    try:
        response = supabase.table("allowed_users").select("*").order("created_at", desc=True).execute()
        return response.data
    except Exception as e:
        print(f"Error fetching allowed users: {e}")
        # Return empty list if table doesn't exist
        return []

@api_router.post("/allowed-users")
async def add_allowed_user(data: dict):
    """Add email to whitelist"""
    try:
        email = data.get("email", "").lower().strip()
        
        if not email:
            raise HTTPException(status_code=400, detail="E-posta adresi gerekli")
        
        # Check if already exists
        try:
            existing = supabase.table("allowed_users").select("*").eq("email", email).execute()
            if existing.data:
                raise HTTPException(status_code=400, detail="Bu e-posta adresi zaten ekli")
        except Exception:
            pass  # Table might not exist yet
        
        new_user = {
            "id": str(uuid.uuid4()),
            "email": email,
            "role": "user",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        supabase.table("allowed_users").insert(new_user).execute()
        return {"message": "Kullanıcı eklendi", "user": new_user}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error adding allowed user: {e}")
        raise HTTPException(status_code=500, detail=f"Kullanıcı eklenirken hata: {str(e)}")

@api_router.delete("/allowed-users/{user_id}")
async def remove_allowed_user(user_id: str):
    """Remove email from whitelist"""
    try:
        response = supabase.table("allowed_users").delete().eq("id", user_id).execute()
        return {"message": "Kullanıcı silindi"}
    except Exception as e:
        print(f"Error removing allowed user: {e}")
        raise HTTPException(status_code=500, detail=f"Kullanıcı silinirken hata: {str(e)}")

# Emergent Auth Session Endpoint
@api_router.post("/auth/session")
async def auth_session(request: Request, response: Response):
    """Exchange Emergent Auth session_id for user data"""
    data = await request.json()
    session_id = data.get("session_id")
    
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id gerekli")
    
    try:
        # Verify session with Emergent Auth - correct endpoint
        async with httpx.AsyncClient() as client:
            auth_response = await client.get(
                "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
                headers={"X-Session-ID": session_id},
                timeout=10.0
            )
            
            if auth_response.status_code != 200:
                logging.error(f"Emergent Auth returned {auth_response.status_code}: {auth_response.text}")
                raise HTTPException(status_code=401, detail="Geçersiz session")
            
            user_data = auth_response.json()
    except httpx.RequestError as e:
        logging.error(f"Emergent Auth error: {e}")
        raise HTTPException(status_code=500, detail="Auth servisi ile iletişim hatası")
    
    email = user_data.get("email", "")
    name = user_data.get("name", "")
    picture = user_data.get("picture", "")
    user_id = user_data.get("id", f"user_{uuid.uuid4().hex[:12]}")

    # SECURITY: Whitelist check — only approved emails can login via Google.
    # Without this, anyone with a Google account could sign in.
    if not email:
        raise HTTPException(status_code=400, detail="E-posta adresi alınamadı")

    admin_email = "hakanonel05@gmail.com"
    email_lc = email.lower()
    whitelist_response = supabase.table("allowed_users").select("email").eq("email", email_lc).execute()
    is_allowed = (
        email_lc == admin_email.lower()
        or (whitelist_response.data and len(whitelist_response.data) > 0)
    )
    if not is_allowed:
        logging.warning(f"Auth blocked — email not in whitelist: {email_lc}")
        raise HTTPException(
            status_code=403,
            detail="Bu e-posta adresi için erişim izni yok. Lütfen yöneticiyle iletişime geçin."
        )

    role = get_user_role(email)
    
    # Check if user exists in DB
    existing = supabase.table("users").select("*").eq("email", email).execute()
    
    if existing.data:
        user_doc = existing.data[0]
        supabase.table("users").update({
            "name": name,
            "picture": picture
        }).eq("email", email).execute()
    else:
        user_doc = {
            "user_id": user_id,
            "email": email,
            "name": name,
            "picture": picture,
            "role": role,
            "auth_type": "google",
            "notification_days": 3,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        supabase.table("users").insert(user_doc).execute()
    
    # Create session
    session_token = f"session_{uuid.uuid4().hex}"
    session_data = SessionData(
        user_id=user_id,
        email=email,
        name=name,
        picture=picture,
        role=role
    )
    sessions[session_token] = session_data
    _save_sessions_to_disk()
    
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        max_age=7*24*60*60
    )
    
    return {
        "user_id": user_id,
        "email": email,
        "name": name,
        "picture": picture,
        "role": role,
        "session_token": session_token
    }

# Local Auth
class LocalLoginRequest(BaseModel):
    email: str
    password: str

class LocalRegisterRequest(BaseModel):
    email: str
    password: str
    name: str

@api_router.post("/auth/register")
async def register(data: LocalRegisterRequest, response: Response):
    existing = supabase.table("users").select("*").eq("email", data.email).execute()
    if existing.data:
        raise HTTPException(status_code=400, detail="Bu e-posta adresi zaten kayıtlı")
    
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    role = get_user_role(data.email)
    password_hash = pwd_context.hash(data.password)
    
    user_doc = {
        "user_id": user_id,
        "email": data.email,
        "name": data.name,
        "picture": "",
        "role": role,
        "auth_type": "local",
        "password_hash": password_hash,
        "notification_days": 3,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    supabase.table("users").insert(user_doc).execute()
    
    session_token = f"session_{uuid.uuid4().hex}"
    session_data = SessionData(
        user_id=user_id,
        email=data.email,
        name=data.name,
        picture="",
        role=role
    )
    sessions[session_token] = session_data
    _save_sessions_to_disk()
    
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        max_age=7*24*60*60
    )
    
    return {
        "user_id": user_id,
        "email": data.email,
        "name": data.name,
        "picture": "",
        "role": role,
        "auth_type": "local",
        "notification_days": 3,
        "created_at": user_doc["created_at"],
        "session_token": session_token
    }

@api_router.post("/auth/login")
async def login(data: LocalLoginRequest, response: Response):
    # First check whitelist
    email = data.email.lower().strip()
    admin_email = "hakanonel05@gmail.com"
    
    try:
        whitelist_response = supabase.table("allowed_users").select("*").eq("email", email).execute()
        is_allowed = email == admin_email.lower() or (whitelist_response.data and len(whitelist_response.data) > 0)
    except Exception:
        is_allowed = email == admin_email.lower()
    
    if not is_allowed:
        raise HTTPException(
            status_code=403, 
            detail="Bu e-posta adresi için erişim izni yok. Lütfen yöneticiyle iletişime geçin."
        )
    
    existing = supabase.table("users").select("*").eq("email", data.email).execute()
    if not existing.data:
        raise HTTPException(status_code=401, detail="Geçersiz e-posta veya şifre")
    
    user = existing.data[0]
    
    if user.get("auth_type") == "google":
        raise HTTPException(status_code=400, detail="Bu hesap Google ile giriş yapmalıdır")
    
    if not pwd_context.verify(data.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Geçersiz e-posta veya şifre")
    
    session_token = f"session_{uuid.uuid4().hex}"
    session_data = SessionData(
        user_id=user["user_id"],
        email=user["email"],
        name=user["name"],
        picture=user.get("picture", ""),
        role=user["role"]
    )
    sessions[session_token] = session_data
    _save_sessions_to_disk()
    
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        max_age=7*24*60*60
    )
    
    return {
        "user_id": user["user_id"],
        "email": user["email"],
        "name": user["name"],
        "picture": user.get("picture", ""),
        "role": user["role"],
        "auth_type": user["auth_type"],
        "notification_days": user.get("notification_days", 3),
        "created_at": user["created_at"],
        "session_token": session_token
    }

# ============ USERS MANAGEMENT ============

@api_router.get("/users")
async def get_users(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user_from_request(request, session_token)
    if not check_super_admin(user):
        raise HTTPException(status_code=403, detail="Bu işlem için süper admin yetkisi gerekli")
    
    response = supabase.table("users").select("user_id, email, name, picture, role, auth_type, notification_days, created_at").execute()
    data = response.data or []
    # Attach effective permissions for each user (admin → all true)
    for u in data:
        if (u.get("email") or "").lower() == ADMIN_EMAIL.lower() or u.get("role") == "admin":
            u["permissions"] = {k: True for k in ALL_PERMISSION_KEYS}
            u["is_admin"] = True
        else:
            u["permissions"] = get_user_permissions(u.get("user_id"))
            u["is_admin"] = False
    return data

@api_router.put("/users/{user_id}/role")
@api_router.patch("/users/{user_id}/role")
async def update_user_role(
    user_id: str,
    request: Request,
    role: Optional[str] = None,
    role_data: Optional[dict] = Body(default=None),
    session_token: Optional[str] = Cookie(None),
):
    user = await get_current_user_from_request(request, session_token)
    if not check_super_admin(user):
        raise HTTPException(status_code=403, detail="Bu işlem için süper admin yetkisi gerekli")
    
    # Accept both query param `?role=...` and body `{"role": "..."}`
    new_role = role or (role_data or {}).get("role") or "user"
    if new_role not in ["admin", "user"]:
        raise HTTPException(status_code=400, detail="Geçersiz rol")
    
    supabase.table("users").update({"role": new_role}).eq("user_id", user_id).execute()
    return {"message": "Rol güncellendi", "role": new_role}


@api_router.get("/users/{user_id}/permissions")
async def get_user_permissions_endpoint(
    user_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None),
):
    user = await get_current_user_from_request(request, session_token)
    if not check_super_admin(user):
        raise HTTPException(status_code=403, detail="Bu işlem için süper admin yetkisi gerekli")
    return {"user_id": user_id, "permissions": get_user_permissions(user_id)}


@api_router.patch("/users/{user_id}/permissions")
async def update_user_permissions(
    user_id: str,
    request: Request,
    payload: dict = Body(...),
    session_token: Optional[str] = Cookie(None),
):
    """
    Body: { "can_delete": true, "can_edit_dashboard": false }
    Only the included keys are updated. Unknown keys are ignored.
    """
    user = await get_current_user_from_request(request, session_token)
    if not check_super_admin(user):
        raise HTTPException(status_code=403, detail="Bu işlem için süper admin yetkisi gerekli")
    for key, value in (payload or {}).items():
        if key in ALL_PERMISSION_KEYS:
            set_user_permission(user_id, key, value)
    return {"user_id": user_id, "permissions": get_user_permissions(user_id)}


@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user_from_request(request, session_token)
    if not check_super_admin(user):
        raise HTTPException(status_code=403, detail="Bu işlem için süper admin yetkisi gerekli")
    
    supabase.table("users").delete().eq("user_id", user_id).execute()
    return {"message": "Kullanıcı silindi"}

@api_router.delete("/users/cleanup/unauthorized")
async def cleanup_unauthorized_users(request: Request, session_token: Optional[str] = Cookie(None)):
    """Delete all users that are not in the allowed_users whitelist AND revoke their active sessions."""
    user = await get_current_user_from_request(request, session_token)
    if not check_super_admin(user):
        raise HTTPException(status_code=403, detail="Bu işlem için süper admin yetkisi gerekli")
    
    # Get allowed emails
    allowed_response = supabase.table("allowed_users").select("email").execute()
    allowed_emails = [u["email"].lower() for u in allowed_response.data]
    
    # Always keep admin
    admin_email = "hakanonel05@gmail.com"
    if admin_email.lower() not in allowed_emails:
        allowed_emails.append(admin_email.lower())
    
    # Get all users
    users_response = supabase.table("users").select("user_id, email").execute()
    
    deleted_count = 0
    deleted_emails = []
    unauthorized_emails_lc = set()
    
    for u in users_response.data:
        if u["email"].lower() not in allowed_emails:
            unauthorized_emails_lc.add(u["email"].lower())
            supabase.table("users").delete().eq("user_id", u["user_id"]).execute()
            deleted_count += 1
            deleted_emails.append(u["email"])

    # SECURITY: Also revoke active sessions for the unauthorized users.
    # Without this, an unauthorized user who already logged in stays logged in
    # until their session expires (7 days).
    revoked_sessions = 0
    if unauthorized_emails_lc:
        # In-memory sessions
        for token in list(sessions.keys()):
            sess = sessions.get(token)
            sess_email = (getattr(sess, "email", "") or "").lower()
            if sess_email in unauthorized_emails_lc:
                del sessions[token]
                _delete_persisted_session(token)
                revoked_sessions += 1
        # DB-persisted sessions (in case in-memory cache hasn't loaded them yet)
        try:
            for em in unauthorized_emails_lc:
                supabase.table("sessions").delete().eq("email", em).execute()
        except Exception as e:
            logging.warning(f"Could not purge DB sessions for unauthorized users: {e}")
        _save_sessions_to_disk()

    return {
        "message": f"{deleted_count} yetkisiz kullanıcı silindi, {revoked_sessions} aktif oturum iptal edildi",
        "deleted_emails": deleted_emails,
        "revoked_sessions": revoked_sessions,
    }

# ============ NOTIFICATIONS ============

@api_router.get("/notifications")
async def get_notifications(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user_from_request(request, session_token)
    
    notification_days = 3
    if user:
        user_doc = supabase.table("users").select("notification_days").eq("email", user["email"]).execute()
        if user_doc.data:
            notification_days = user_doc.data[0].get("notification_days", 3)
    
    threshold_date = (datetime.now(timezone.utc) + timedelta(days=notification_days)).strftime("%Y-%m-%d")
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Get customers with upcoming followups
    customers = supabase.table("customers").select("*").execute().data
    followup_customers = [c for c in customers if c.get("next_followup_date") and c["next_followup_date"] <= threshold_date and c["next_followup_date"] >= today]
    
    # Get visits with upcoming dates
    visits = supabase.table("visits").select("*").execute().data
    upcoming_visits = [v for v in visits if v.get("next_visit_date") and v["next_visit_date"] <= threshold_date and v["next_visit_date"] >= today]
    
    notifications = []
    
    for customer in followup_customers:
        notifications.append({
            "id": f"followup_{customer['id']}",
            "type": "followup",
            "title": f"Takip Hatırlatması: {customer['company_name']}",
            "message": f"Takip tarihi: {customer['next_followup_date']}",
            "date": customer['next_followup_date'],
            "customer_id": customer['id'],
            "customer_name": customer['company_name']
        })
    
    for visit in upcoming_visits:
        customer = supabase.table("customers").select("company_name").eq("id", visit["customer_id"]).execute()
        customer_name = customer.data[0]["company_name"] if customer.data else "Bilinmiyor"
        
        notifications.append({
            "id": f"visit_{visit['id']}",
            "type": "visit",
            "title": f"Yaklaşan Ziyaret: {customer_name}",
            "message": f"Ziyaret tarihi: {visit['next_visit_date']}",
            "date": visit['next_visit_date'],
            "customer_id": visit['customer_id'],
            "customer_name": customer_name
        })
    
    notifications.sort(key=lambda x: x['date'])
    
    return {
        "notifications": notifications,
        "count": len(notifications),
        "notification_days": notification_days
    }

@api_router.put("/settings/notifications")
async def update_notification_settings(settings: dict, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user_from_request(request, session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Oturum açılmamış")
    
    days = settings.get("notification_days", 3)
    supabase.table("users").update({"notification_days": days}).eq("email", user["email"]).execute()
    
    return {"message": "Ayarlar güncellendi", "notification_days": days}

# ============ CALENDAR EVENTS ============

@api_router.get("/calendar-events")
async def get_calendar_events():
    visits = supabase.table("visits").select("*").execute().data
    customers = supabase.table("customers").select("id, company_name, next_followup_date").execute().data
    
    events = []
    
    for visit in visits:
        if visit.get("visit_date"):
            customer = next((c for c in customers if c["id"] == visit["customer_id"]), None)
            customer_name = customer["company_name"] if customer else "Bilinmiyor"
            
            events.append({
                "id": f"visit_{visit['id']}",
                "title": f"Ziyaret: {customer_name}",
                "start": visit["visit_date"],
                "type": "visit",
                "customer_id": visit["customer_id"],
                "visit_type": visit.get("visit_type", ""),
                "color": "#10B981"
            })
        
        if visit.get("next_visit_date"):
            customer = next((c for c in customers if c["id"] == visit["customer_id"]), None)
            customer_name = customer["company_name"] if customer else "Bilinmiyor"
            
            events.append({
                "id": f"next_visit_{visit['id']}",
                "title": f"Planl Ziyaret: {customer_name}",
                "start": visit["next_visit_date"],
                "type": "planned_visit",
                "customer_id": visit["customer_id"],
                "color": "#3B82F6"
            })
    
    for customer in customers:
        if customer.get("next_followup_date"):
            events.append({
                "id": f"followup_{customer['id']}",
                "title": f"Takip: {customer['company_name']}",
                "start": customer["next_followup_date"],
                "type": "followup",
                "customer_id": customer["id"],
                "color": "#F59E0B"
            })
    
    return events

# ============ KANBAN ENDPOINTS ============

KANBAN_STATUSES = [
    {"id": "beklemede", "title": "Beklemede", "color": "amber"},
    {"id": "iletisimde", "title": "İletişimde", "color": "blue"},
    {"id": "teklif_verildi", "title": "Teklif Verildi", "color": "purple"},
    {"id": "calisiliyor", "title": "Çalışılıyor", "color": "emerald"},
    {"id": "kazanildi", "title": "Kazanıldı", "color": "green"},
    {"id": "kaybedildi", "title": "Kaybedildi", "color": "red"},
]

KANBAN_GROUP_FIELDS = {
    "status": {
        "label": "Durum",
        "default_columns": ["Beklemede", "İletişimde", "Teklif Verildi", "Çalışılıyor", "Kazanıldı", "Kaybedildi"],
        "colors": {
            "Beklemede": "amber", "İletişimde": "blue", "Teklif Verildi": "purple",
            "Çalışılıyor": "emerald", "Kazanıldı": "green", "Kaybedildi": "red"
        }
    },
    "assigned_to": {
        "label": "Takip Eden", 
        "default_columns": ["Hakan ÖNEL", "Furkan Çelik", "Melih Karaman"], 
        "colors": {
            "Hakan ÖNEL": "emerald",
            "Furkan Çelik": "blue",
            "Melih Karaman": "purple"
        }
    },
    "potential_level": {
        "label": "Potansiyel",
        "default_columns": ["Yüksek", "Orta", "Düşük"],
        "colors": {"Yüksek": "emerald", "Orta": "amber", "Düşük": "slate"}
    },
    "city": {"label": "Şehir", "default_columns": [], "colors": {}},
    "market": {"label": "Market", "default_columns": [], "colors": {}},
    "application": {"label": "Uygulama", "default_columns": [], "colors": {}}
}

def normalize_status(status: str) -> str:
    status_map = {
        "beklemede": "Beklemede",
        "iletisimde": "İletişimde",
        "iletişimde": "İletişimde",
        "teklif_verildi": "Teklif Verildi",
        "teklif verildi": "Teklif Verildi",
        "calisiliyor": "Çalışılıyor",
        "çalışılıyor": "Çalışılıyor",
        "kazanildi": "Kazanıldı",
        "kazanıldı": "Kazanıldı",
        "kaybedildi": "Kaybedildi",
        "olumsuz": "Kaybedildi",
        "takip ediliyor": "İletişimde",
        "direkt takip": "Çalışılıyor",
        "fiyat çalışması yapılıyor": "Teklif Verildi",
        "uygun ürün bekleniyor": "Beklemede",
    }
    return status_map.get(status.lower().strip(), status) if status else "Beklemede"

DYNAMIC_COLORS = ["blue", "emerald", "purple", "amber", "rose", "cyan", "indigo", "orange", "teal", "pink"]

@api_router.get("/kanban/statuses")
async def get_kanban_statuses():
    return KANBAN_STATUSES

@api_router.post("/migrate/normalize-statuses")
async def migrate_normalize_statuses():
    """One-time migration to normalize all customer statuses to match Kanban columns"""
    valid_statuses = {"Beklemede", "İletişimde", "Teklif Verildi", "Çalışılıyor", "Kazanıldı", "Kaybedildi"}
    
    # Fetch all customers
    all_customers = []
    page_size = 1000
    offset = 0
    while True:
        response = supabase.table("customers").select("id, status").range(offset, offset + page_size - 1).execute()
        if not response.data:
            break
        all_customers.extend(response.data)
        if len(response.data) < page_size:
            break
        offset += page_size
    
    updated = 0
    for customer in all_customers:
        old_status = customer.get("status", "")
        if old_status and old_status not in valid_statuses:
            new_status = normalize_status(old_status)
            if new_status != old_status:
                supabase.table("customers").update({
                    "status": new_status,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }).eq("id", customer["id"]).execute()
                updated += 1
    
    return {"message": f"{updated} müşterinin durumu güncellendi", "updated": updated}

@api_router.get("/kanban/group-fields")
async def get_kanban_group_fields():
    return [{"value": key, "label": val["label"]} for key, val in KANBAN_GROUP_FIELDS.items()]

@api_router.get("/kanban/customers")
async def get_kanban_customers(group_by: str = "status"):
    # In-memory cache (30s TTL) - same group_by gets cached response
    now = _time.time()
    cached = _kanban_cache.get(group_by)
    if cached and (now - cached["timestamp"]) < _KANBAN_CACHE_TTL:
        return cached["data"]
    
    # Only select fields needed for Kanban cards
    kanban_fields = "id, company_name, market, application, city, status, potential_level, assigned_to, contact_info, products, competitor, partner"
    
    all_customers = []
    page_size = 1000
    offset = 0
    while True:
        response = supabase.table("customers").select(kanban_fields).range(offset, offset + page_size - 1).execute()
        if not response.data:
            break
        all_customers.extend(response.data)
        if len(response.data) < page_size:
            break
        offset += page_size
    
    customers = all_customers
    
    field_config = KANBAN_GROUP_FIELDS.get(group_by, KANBAN_GROUP_FIELDS["status"])
    
    if field_config["default_columns"]:
        columns = field_config["default_columns"].copy()
    else:
        unique_values = set()
        for customer in customers:
            value = customer.get(group_by, "")
            if value:
                unique_values.add(value)
        columns = sorted(list(unique_values)) if unique_values else ["Atanmamış"]
    
    if "Atanmamış" not in columns:
        columns.append("Atanmamış")
    
    grouped = {col: [] for col in columns}
    
    for customer in customers:
        value = customer.get(group_by, "")
        
        if group_by == "status":
            value = normalize_status(value) if value else "Beklemede"
        
        if not value:
            value = "Atanmamış"
        
        if value in grouped:
            grouped[value].append(customer)
        elif "Atanmamış" in grouped:
            grouped["Atanmamış"].append(customer)
    
    if "Atanmamış" in grouped and len(grouped["Atanmamış"]) == 0 and len(columns) > 1:
        del grouped["Atanmamış"]
    
    # Cache the result
    _kanban_cache[group_by] = {"data": grouped, "timestamp": now}
    return grouped

@api_router.get("/kanban/columns")
async def get_kanban_columns(group_by: str = "status"):
    field_config = KANBAN_GROUP_FIELDS.get(group_by, KANBAN_GROUP_FIELDS["status"])
    
    if field_config["default_columns"]:
        columns = field_config["default_columns"]
        colors = field_config["colors"]
    else:
        response = supabase.table("customers").select(group_by).execute()
        unique_values = set()
        for customer in response.data:
            value = customer.get(group_by, "")
            if value:
                unique_values.add(value)
        columns = sorted(list(unique_values)) if unique_values else ["Atanmamış"]
        
        colors = {}
        for idx, col in enumerate(columns):
            colors[col] = DYNAMIC_COLORS[idx % len(DYNAMIC_COLORS)]
    
    return {"columns": columns, "colors": colors, "field_label": field_config["label"]}

# Kanban Views CRUD
@api_router.get("/kanban/views", response_model=List[KanbanView])
async def get_kanban_views():
    response = supabase.table("kanban_views").select("*").execute()
    return response.data

@api_router.post("/kanban/views", response_model=KanbanView)
async def create_kanban_view(view_data: KanbanViewCreate):
    view_obj = KanbanView(**view_data.model_dump())
    doc = view_obj.model_dump()
    
    supabase.table("kanban_views").insert(doc).execute()
    return view_obj

@api_router.put("/kanban/views/{view_id}", response_model=KanbanView)
async def update_kanban_view(view_id: str, view_data: KanbanViewCreate):
    response = supabase.table("kanban_views").select("*").eq("id", view_id).execute()
    if not response.data:
        raise HTTPException(status_code=404, detail="Kanban görünümü bulunamadı")
    
    update_data = view_data.model_dump()
    supabase.table("kanban_views").update(update_data).eq("id", view_id).execute()
    
    updated = supabase.table("kanban_views").select("*").eq("id", view_id).execute()
    return updated.data[0]

@api_router.delete("/kanban/views/{view_id}")
async def delete_kanban_view(view_id: str):
    response = supabase.table("kanban_views").select("*").eq("id", view_id).execute()
    if not response.data:
        raise HTTPException(status_code=404, detail="Kanban görünümü bulunamadı")
    
    supabase.table("kanban_views").delete().eq("id", view_id).execute()
    return {"message": "Kanban görünümü silindi"}

@api_router.patch("/kanban/customers/{customer_id}/status")
async def update_customer_status(customer_id: str, new_status: str = Query(...)):
    valid_statuses = [s["title"] for s in KANBAN_STATUSES]
    if new_status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Geçersiz durum. Geçerli durumlar: {', '.join(valid_statuses)}")
    
    supabase.table("customers").update({
        "status": new_status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }).eq("id", customer_id).execute()
    
    _invalidate_kanban_cache()
    _stats_cache["data"] = None
    return {"message": "Durum güncellendi", "new_status": new_status}

@api_router.patch("/kanban/customers/{customer_id}/field")
async def update_customer_field(customer_id: str, field: str = Query(...), value: str = Query(...)):
    valid_fields = list(KANBAN_GROUP_FIELDS.keys())
    if field not in valid_fields:
        raise HTTPException(status_code=400, detail=f"Geçersiz alan. Geçerli alanlar: {', '.join(valid_fields)}")
    
    supabase.table("customers").update({
        field: value,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }).eq("id", customer_id).execute()
    
    _invalidate_kanban_cache()
    return {"message": "Alan güncellendi", "field": field, "value": value}

# ============ EXPORT ENDPOINTS ============

@api_router.get("/export/full-backup")
async def export_full_backup(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user_from_request(request, session_token)
    if not check_admin_permission(user):
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
    
    customers = supabase.table("customers").select("*").execute().data
    visits = supabase.table("visits").select("*").execute().data
    calls = supabase.table("calls").select("*").execute().data
    options = supabase.table("options").select("*").execute().data
    filters = supabase.table("saved_filters").select("*").execute().data
    kanban_views = supabase.table("kanban_views").select("*").execute().data
    
    backup_data = {
        "export_date": datetime.now(timezone.utc).isoformat(),
        "customers": customers,
        "visits": visits,
        "calls": calls,
        "options": options,
        "saved_filters": filters,
        "kanban_views": kanban_views
    }
    
    json_content = json.dumps(backup_data, ensure_ascii=False, indent=2)
    filename = f"crmmaster_full_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    
    return StreamingResponse(
        io.BytesIO(json_content.encode('utf-8')),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============ AUTOMATED BACKUPS ============

def _run_scheduled_backup():
    """Callable used by APScheduler. No args."""
    return backup_service.run_backup_sync(
        supabase=supabase,
        resend_module=resend if RESEND_API_KEY else None,
        sender_email=SENDER_EMAIL,
    )


class BackupConfigUpdate(BaseModel):
    enabled: Optional[bool] = None
    frequency: Optional[str] = None       # "daily" | "weekly"
    hour: Optional[int] = None
    minute: Optional[int] = None
    day_of_week: Optional[str] = None     # mon..sun
    email_enabled: Optional[bool] = None
    email_recipients: Optional[List[str]] = None
    retention_days: Optional[int] = None


@api_router.get("/backups/config")
async def get_backup_config(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user_from_request(request, session_token)
    if not check_admin_permission(user):
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
    cfg = backup_service.load_config()
    cfg["next_run"] = backup_service.next_run_time()
    return cfg


@api_router.put("/backups/config")
async def update_backup_config(payload: BackupConfigUpdate, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user_from_request(request, session_token)
    if not check_admin_permission(user):
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
    updates = {k: v for k, v in payload.model_dump(exclude_unset=True).items() if v is not None}
    # Validate frequency
    if "frequency" in updates and updates["frequency"] not in ("daily", "weekly"):
        raise HTTPException(status_code=400, detail="Frequency must be 'daily' or 'weekly'")
    if "hour" in updates and not (0 <= int(updates["hour"]) <= 23):
        raise HTTPException(status_code=400, detail="Hour must be 0-23")
    if "minute" in updates and not (0 <= int(updates["minute"]) <= 59):
        raise HTTPException(status_code=400, detail="Minute must be 0-59")
    if "day_of_week" in updates and updates["day_of_week"] not in ("mon","tue","wed","thu","fri","sat","sun"):
        raise HTTPException(status_code=400, detail="Invalid day_of_week")
    if "retention_days" in updates and int(updates["retention_days"]) < 1:
        raise HTTPException(status_code=400, detail="Retention must be >= 1")
    backup_service.save_config(updates)
    backup_service.apply_schedule(_run_scheduled_backup)
    cfg = backup_service.load_config()
    cfg["next_run"] = backup_service.next_run_time()
    return cfg


@api_router.get("/backups/list")
async def list_backups_endpoint(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user_from_request(request, session_token)
    if not check_admin_permission(user):
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
    return {"backups": backup_service.list_backups()}


@api_router.post("/backups/trigger")
async def trigger_backup(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user_from_request(request, session_token)
    if not check_admin_permission(user):
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
    result = await asyncio.to_thread(_run_scheduled_backup)
    return result


@api_router.get("/backups/{filename}/download")
async def download_backup(filename: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user_from_request(request, session_token)
    if not check_admin_permission(user):
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
    path = backup_service.get_backup_path(filename)
    if not path:
        raise HTTPException(status_code=404, detail="Yedek dosyası bulunamadı")
    return FileResponse(
        path=str(path),
        media_type="application/json",
        filename=filename,
    )


@api_router.delete("/backups/{filename}")
async def delete_backup_endpoint(filename: str, request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user_from_request(request, session_token)
    if not check_admin_permission(user):
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
    ok = backup_service.delete_backup(filename)
    if not ok:
        raise HTTPException(status_code=404, detail="Yedek dosyası bulunamadı")
    return {"ok": True}


@api_router.get("/export/customers/xlsx")
async def export_customers_xlsx(request: Request, session_token: Optional[str] = Cookie(None)):
    user = await get_current_user_from_request(request, session_token)
    if not check_admin_permission(user):
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
    
    response = supabase.table("customers").select("*").execute()
    customers = response.data
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Müşteriler"
    
    headers = ['Firma Adı', 'Market', 'Uygulama', 'Şehir', 'İlçe', 'Web', 'Durum',
               'İletişim Kişisi', 'E-posta', 'Telefon', 'Rakip', 'Partner',
               'Potansiyel Seviye', 'Takip Eden', 'ABB Ürünleri', 'Notlar']
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, customer in enumerate(customers, 2):
        contact = customer.get("contact_info", {}) or {}
        products = ",".join(customer.get("products", []) or [])
        
        row_data = [
            customer.get("company_name", ""),
            customer.get("market", ""),
            customer.get("application", ""),
            customer.get("city", ""),
            customer.get("district", ""),
            customer.get("website", ""),
            customer.get("status", ""),
            contact.get("contact_person", ""),
            contact.get("email", ""),
            contact.get("phone", ""),
            customer.get("competitor", ""),
            customer.get("partner", ""),
            customer.get("potential_level", ""),
            customer.get("assigned_to", ""),
            products,
            customer.get("notes", "")
        ]
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"crm_musteri_yedek_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/export/visits")
async def export_visits_xlsx(request: Request, session_token: Optional[str] = Cookie(None)):
    """Export all visits as XLSX - Admin only"""
    user = await get_current_user_from_request(request, session_token)
    if not check_admin_permission(user):
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
    
    response = supabase.table("visits").select("*").execute()
    visits = response.data
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ziyaretler"
    
    headers = ['Tarih', 'Firma', 'Ziyaret Eden', 'Ziyaret Tipi', 'Görüşülen Kişi', 
               'Sonuç', 'Sonraki Ziyaret', 'Notlar']
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    for row_idx, visit in enumerate(visits, 2):
        ws.cell(row=row_idx, column=1, value=visit.get('date', ''))
        ws.cell(row=row_idx, column=2, value=visit.get('customer_name', ''))
        ws.cell(row=row_idx, column=3, value=visit.get('visited_by', ''))
        ws.cell(row=row_idx, column=4, value=visit.get('visit_type', ''))
        ws.cell(row=row_idx, column=5, value=visit.get('contact_person', ''))
        ws.cell(row=row_idx, column=6, value=visit.get('outcome', ''))
        ws.cell(row=row_idx, column=7, value=visit.get('next_visit_date', ''))
        ws.cell(row=row_idx, column=8, value=visit.get('notes', ''))
    
    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"crm_ziyaret_yedek_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============ CUSTOM REPORTS ============

class ReportRequest(BaseModel):
    report_type: str = "customers"
    customer_columns: List[str] = []
    contact_columns: List[str] = []
    visit_columns: List[str] = []
    call_columns: List[str] = []
    include_notes: bool = False
    customer_limit: Optional[int] = None
    filters: dict = {}

@api_router.post("/reports/generate")
async def generate_custom_report(data: ReportRequest, request: Request, session_token: Optional[str] = Cookie(None)):
    """Generate custom Excel report with selected columns"""
    user = await get_current_user_from_request(request, session_token)
    if not user:
        raise HTTPException(
            status_code=401,
            detail="Oturum süreniz dolmuş. Lütfen tekrar giriş yapın."
        )
    if not check_admin_permission(user):
        raise HTTPException(
            status_code=403,
            detail="Bu işlem için admin yetkisi gerekli. Sadece yöneticiler rapor oluşturabilir."
        )
    
    # Get ALL customers with filters (Supabase has 1000 row limit, need pagination)
    async def fetch_all_customers():
        all_customers = []
        page_size = 1000
        offset = 0
        
        while True:
            query = supabase.table("customers").select("*")
            
            filters = data.filters or {}
            if filters.get("status"):
                query = query.eq("status", filters["status"])
            if filters.get("is_followup") is not None:
                query = query.eq("is_followup", filters["is_followup"])
            
            # Add pagination
            query = query.range(offset, offset + page_size - 1)
            response = query.execute()
            
            if not response.data:
                break
                
            all_customers.extend(response.data)
            
            if len(response.data) < page_size:
                break
                
            offset += page_size
        
        return all_customers
    
    customers = await fetch_all_customers()
    
    # Apply assigned_to filter case-insensitive in Python
    filters = data.filters or {}
    if filters.get("assigned_to"):
        assigned_lower = filters["assigned_to"].lower()
        customers = [c for c in customers if (c.get("assigned_to") or "").lower() == assigned_lower]
    
    # Apply customer limit if specified
    if data.customer_limit and data.customer_limit > 0:
        customers = customers[:data.customer_limit]
    
    # Get visits and calls if needed (also with pagination)
    visits_dict = {}
    calls_dict = {}
    
    if data.visit_columns:
        all_visits = []
        offset = 0
        while True:
            visits_response = supabase.table("visits").select("*").range(offset, offset + 999).execute()
            if not visits_response.data:
                break
            all_visits.extend(visits_response.data)
            if len(visits_response.data) < 1000:
                break
            offset += 1000
        
        for v in all_visits:
            cid = v.get("customer_id")
            if cid not in visits_dict:
                visits_dict[cid] = []
            visits_dict[cid].append(v)
    
    if data.call_columns:
        all_calls = []
        offset = 0
        while True:
            calls_response = supabase.table("calls").select("*").range(offset, offset + 999).execute()
            if not calls_response.data:
                break
            all_calls.extend(calls_response.data)
            if len(calls_response.data) < 1000:
                break
            offset += 1000
        
        for c in all_calls:
            cid = c.get("customer_id")
            if cid not in calls_dict:
                calls_dict[cid] = []
            calls_dict[cid].append(c)
    
    # Column mappings
    customer_col_map = {
        "company_name": "Firma Adı",
        "market": "Market",
        "application": "Uygulama",
        "city": "Şehir",
        "district": "İlçe",
        "web": "Web Sitesi",
        "website": "Web Sitesi",
        "status": "Durum",
        "potential_level": "Potansiyel",
        "assigned_to": "Takip Eden",
        "competitor": "Rakip",
        "partner": "Partner",
        "products": "Ürünler",
        "is_followup": "Takipte Mi?",
        "next_followup_date": "Sonraki Takip",
        "created_at": "Kayıt Tarihi"
    }
    
    contact_col_map = {
        "contact_name": "Kişi Adı",
        "contact_title": "Unvan",
        "contact_phone": "Telefon",
        "contact_email": "E-posta"
    }
    
    visit_col_map = {
        "visit_date": "Ziyaret Tarihi",
        "visit_type": "Ziyaret Tipi",
        "visited_by": "Ziyaret Eden",
        "visit_outcome": "Sonuç",
        "visit_notes": "Ziyaret Notları"
    }
    
    call_col_map = {
        "call_date": "Arama Tarihi",
        "call_status": "Arama Durumu",
        "call_notes": "Arama Notları"
    }
    
    # Build headers
    headers = []
    for col in data.customer_columns:
        headers.append(customer_col_map.get(col, col))
    for col in data.contact_columns:
        headers.append(contact_col_map.get(col, col))
    for col in data.visit_columns:
        headers.append(visit_col_map.get(col, col))
    for col in data.call_columns:
        headers.append(call_col_map.get(col, col))
    if data.include_notes:
        headers.append("Notlar")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapor"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Build rows
    row_idx = 2
    for customer in customers:
        # Get contact info
        contacts = customer.get("contact_info", {})
        if isinstance(contacts, dict):
            contacts = [contacts] if contacts.get("name") else []
        elif not isinstance(contacts, list):
            contacts = []
        
        # Get customer visits and calls
        cust_visits = visits_dict.get(customer.get("id"), [])
        cust_calls = calls_dict.get(customer.get("id"), [])
        
        # Determine how many rows we need
        max_rows = max(1, len(contacts) if data.contact_columns else 1, 
                       len(cust_visits) if data.visit_columns else 1,
                       len(cust_calls) if data.call_columns else 1)
        
        for i in range(max_rows):
            col_idx = 1
            
            # Customer columns (only on first row or repeated)
            for col in data.customer_columns:
                # Map "web" to "website" for database field
                db_col = "website" if col == "web" else col
                value = customer.get(db_col, "")
                if col == "is_followup":
                    value = "Evet" if value else "Hayır"
                elif col == "created_at" and value:
                    value = value[:10]
                ws.cell(row=row_idx, column=col_idx, value=value if i == 0 else "")
                col_idx += 1
            
            # Contact columns
            for col in data.contact_columns:
                contact = contacts[i] if i < len(contacts) else {}
                if col == "contact_name":
                    value = contact.get("name", "")
                elif col == "contact_title":
                    value = contact.get("title", "")
                elif col == "contact_phone":
                    value = contact.get("phone", "")
                elif col == "contact_email":
                    value = contact.get("email", "")
                else:
                    value = ""
                ws.cell(row=row_idx, column=col_idx, value=value)
                col_idx += 1
            
            # Visit columns
            for col in data.visit_columns:
                visit = cust_visits[i] if i < len(cust_visits) else {}
                if col == "visit_date":
                    value = visit.get("date", "")
                elif col == "visit_type":
                    value = visit.get("visit_type", "")
                elif col == "visited_by":
                    value = visit.get("visited_by", "")
                elif col == "visit_outcome":
                    value = visit.get("outcome", "")
                elif col == "visit_notes":
                    value = visit.get("notes", "")
                else:
                    value = ""
                ws.cell(row=row_idx, column=col_idx, value=value)
                col_idx += 1
            
            # Call columns
            for col in data.call_columns:
                call = cust_calls[i] if i < len(cust_calls) else {}
                if col == "call_date":
                    value = call.get("call_date", "")
                elif col == "call_status":
                    value = call.get("status", "")
                elif col == "call_notes":
                    value = call.get("notes", "")
                else:
                    value = ""
                ws.cell(row=row_idx, column=col_idx, value=value)
                col_idx += 1
            
            # Notes
            if data.include_notes:
                notes_list = customer.get("notes_list", [])
                notes_text = "; ".join([n.get("text", "") for n in notes_list if isinstance(n, dict)]) if i == 0 else ""
                ws.cell(row=row_idx, column=col_idx, value=notes_text)
            
            row_idx += 1
    
    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value or "")) > max_length:
                    max_length = len(str(cell.value or ""))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"rapor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============ TEMPLATE DOWNLOAD ENDPOINTS ============

@api_router.get("/export/customers/template")
async def download_customer_template():
    """Download XLSX template for customer import with all columns"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Şirket Şablonu"
    
    # All customer columns
    headers = [
        'Firma Adı', 'Market', 'Uygulama', 'Şehir', 'İlçe', 'Web Site', 
        'Durum', 'İletişim Kişisi', 'E-posta', 'Telefon', 'Rakip', 'Partner',
        'Potansiyel', 'Takip Eden', 'ABB Ürünleri', 'Notlar'
    ]
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Example row
    example_data = [
        'ABC Şirketi A.Ş.', 'Otomotiv', 'CNC', 'İstanbul', 'Kadıköy', 'www.abcsirketi.com',
        'Beklemede', 'Ahmet Yılmaz', 'ahmet@abcsirketi.com', '0532 123 45 67', 'Siemens', 'XYZ Partner',
        'Yüksek', 'Mehmet Demir', 'PLC, Motor', 'Potansiyel müşteri'
    ]
    
    for col, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col, value=value)
    
    # Column widths
    column_widths = [20, 15, 15, 12, 12, 25, 12, 18, 25, 18, 12, 15, 12, 15, 20, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = width
    
    # Add instructions sheet
    ws2 = wb.create_sheet("Talimatlar")
    instructions = [
        ["ŞIRKET İÇE AKTARMA TALİMATLARI", ""],
        ["", ""],
        ["Zorunlu Alan:", "Firma Adı"],
        ["", ""],
        ["Durum Değerleri:", "Beklemede, İletişimde, Çalışılıyor, Takip Ediliyor, Olumsuz, Kapandı"],
        ["Potansiyel Değerleri:", "Düşük, Orta, Yüksek"],
        ["", ""],
        ["Notlar:", ""],
        ["- İlk satır başlık satırıdır, silmeyin", ""],
        ["- 2. satırdaki örnek veriyi silin veya üzerine yazın", ""],
        ["- ABB Ürünleri sütununda birden fazla ürün virgülle ayrılır", ""],
        ["- Web Site http:// veya https:// olmadan yazılabilir", ""],
    ]
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 60
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sirket_sablonu.xlsx"}
    )

@api_router.get("/export/visits/template")
async def download_visit_template():
    """Download XLSX template for visit import"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ziyaret Şablonu"
    
    headers = ['Tarih', 'Ziyaret Eden', 'Ziyaret Tipi', 'Görüşülen Kişi', 'Sonuç', 'Sonraki Ziyaret', 'Notlar']
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Example row
    example_data = ['2024-01-15', 'Mehmet Demir', 'Yüz Yüze', 'Ahmet Yılmaz', 'Olumlu', '2024-02-15', 'İlk görüşme yapıldı']
    for col, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col, value=value)
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    # Instructions sheet
    ws2 = wb.create_sheet("Talimatlar")
    instructions = [
        ["ZİYARET İÇE AKTARMA TALİMATLARI", ""],
        ["", ""],
        ["Zorunlu Alan:", "Tarih"],
        ["", ""],
        ["Tarih Formatı:", "YYYY-MM-DD (örn: 2024-01-15)"],
        ["Ziyaret Eden:", "Ziyareti yapan kişinin adı"],
        ["Ziyaret Tipi:", "Yüz Yüze, Online, Telefon"],
        ["Görüşülen Kişi:", "Müşteri tarafındaki görüşülen kişi"],
        ["Sonuç:", "Olumlu, Olumsuz, Beklemede, Takip Gerekli"],
        ["", ""],
        ["Not:", "Ziyaretler seçilen şirkete atanacaktır"],
    ]
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 50
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ziyaret_sablonu.xlsx"}
    )

# ============ EMAIL NOTIFICATION ENDPOINTS ============

class NotificationRecipient(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: str
    name: str = ""
    active: bool = True
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class NotificationSettings(BaseModel):
    recipients: List[NotificationRecipient] = []
    days_before: int = 1
    enabled: bool = True

@api_router.get("/notification-settings")
async def get_notification_settings():
    """Get email notification settings"""
    try:
        response = supabase.table("notification_settings").select("*").execute()
        if response.data:
            return response.data[0]
        # Return default settings if none exist
        return {
            "id": "default",
            "recipients": [],
            "days_before": 1,
            "enabled": True
        }
    except Exception as e:
        logging.error(f"Error getting notification settings: {e}")
        return {
            "id": "default",
            "recipients": [],
            "days_before": 1,
            "enabled": True
        }

@api_router.post("/notification-settings")
async def save_notification_settings(settings: dict):
    """Save email notification settings"""
    try:
        # Check if settings exist
        existing = supabase.table("notification_settings").select("*").execute()
        
        if existing.data:
            # Update existing
            supabase.table("notification_settings").update(settings).eq("id", existing.data[0]["id"]).execute()
        else:
            # Create new
            settings["id"] = "default"
            supabase.table("notification_settings").insert(settings).execute()
        
        return {"message": "Ayarlar kaydedildi", "settings": settings}
    except Exception as e:
        logging.error(f"Error saving notification settings: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/notification-settings/add-recipient")
async def add_notification_recipient(recipient: dict):
    """Add a new email recipient"""
    try:
        settings = await get_notification_settings()
        recipients = settings.get("recipients", [])
        
        # Check if email already exists
        if any(r.get("email") == recipient.get("email") for r in recipients):
            raise HTTPException(status_code=400, detail="Bu e-posta adresi zaten ekli")
        
        new_recipient = {
            "id": str(uuid.uuid4()),
            "email": recipient.get("email"),
            "name": recipient.get("name", ""),
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        recipients.append(new_recipient)
        
        settings["recipients"] = recipients
        await save_notification_settings(settings)
        
        return {"message": "Alıcı eklendi", "recipient": new_recipient}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error adding recipient: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/notification-settings/recipient/{recipient_id}")
async def remove_notification_recipient(recipient_id: str):
    """Remove an email recipient"""
    try:
        settings = await get_notification_settings()
        recipients = settings.get("recipients", [])
        
        recipients = [r for r in recipients if r.get("id") != recipient_id]
        settings["recipients"] = recipients
        
        await save_notification_settings(settings)
        
        return {"message": "Alıcı silindi"}
    except Exception as e:
        logging.error(f"Error removing recipient: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/send-followup-reminders")
async def send_followup_reminders():
    """Send email reminders for upcoming follow-ups (1 day before)"""
    try:
        # Get notification settings
        settings = await get_notification_settings()
        
        if not settings.get("enabled", True):
            return {"message": "Bildirimler devre dışı", "sent": 0}
        
        # Resend free plan limitation: only send to account owner email
        PRIMARY_EMAIL = "hakanonel05@gmail.com"
        
        days_before = settings.get("days_before", 1)
        
        # Calculate target date (tomorrow or days_before)
        target_date = (datetime.now(timezone.utc) + timedelta(days=days_before)).strftime("%Y-%m-%d")
        
        # Get customers with follow-up on target date
        customers_response = supabase.table("customers").select("*").eq("next_followup_date", target_date).execute()
        customers = customers_response.data
        
        if not customers:
            return {"message": "Yaklaşan takip yok", "sent": 0}
        
        # Build email content
        customer_list = ""
        for c in customers:
            customer_list += f"""
            <tr>
                <td style="padding: 10px; border-bottom: 1px solid #eee;">{c.get('company_name', '-')}</td>
                <td style="padding: 10px; border-bottom: 1px solid #eee;">{c.get('status', '-')}</td>
                <td style="padding: 10px; border-bottom: 1px solid #eee;">{c.get('assigned_to', '-')}</td>
                <td style="padding: 10px; border-bottom: 1px solid #eee;">{c.get('next_followup_date', '-')}</td>
            </tr>
            """
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
        </head>
        <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
            <div style="background: linear-gradient(135deg, #10b981, #059669); padding: 20px; border-radius: 10px 10px 0 0;">
                <h1 style="color: white; margin: 0;">📅 CRMaster - Takip Hatırlatması</h1>
            </div>
            <div style="background: #f8fafc; padding: 20px; border: 1px solid #e2e8f0; border-radius: 0 0 10px 10px;">
                <p style="color: #334155; font-size: 16px;">
                    Merhaba,<br><br>
                    Yarın ({target_date}) takip edilmesi gereken <strong>{len(customers)}</strong> müşteri var:
                </p>
                
                <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
                    <thead>
                        <tr style="background: #10b981; color: white;">
                            <th style="padding: 10px; text-align: left;">Firma</th>
                            <th style="padding: 10px; text-align: left;">Durum</th>
                            <th style="padding: 10px; text-align: left;">Takip Eden</th>
                            <th style="padding: 10px; text-align: left;">Tarih</th>
                        </tr>
                    </thead>
                    <tbody>
                        {customer_list}
                    </tbody>
                </table>
                
                <p style="color: #64748b; font-size: 14px;">
                    Bu otomatik bir hatırlatma e-postasıdır.<br>
                    CRMaster Müşteri Yönetim Sistemi
                </p>
            </div>
        </body>
        </html>
        """
        
        # Send email only to primary email (Resend free plan limitation)
        try:
            params = {
                "from": SENDER_EMAIL,
                "to": [PRIMARY_EMAIL],
                "subject": f"🔔 CRMaster: {len(customers)} Müşteri Takibi Yarın ({target_date})",
                "html": html_content
            }
            
            await asyncio.to_thread(resend.Emails.send, params)
            logging.info(f"Reminder sent to {PRIMARY_EMAIL}")
            
            # Log activity
            await log_activity(
                activity_type="email_sent",
                title=f"Takip hatırlatma e-postası gönderildi",
                subtitle=f"{len(customers)} müşteri"
            )
            
            return {
                "message": f"E-posta {PRIMARY_EMAIL} adresine gönderildi",
                "sent": 1,
                "customers_count": len(customers)
            }
        except Exception as e:
            logging.error(f"Failed to send reminder: {e}")
            raise HTTPException(status_code=500, detail=f"E-posta gönderilemedi: {str(e)}")
    except Exception as e:
        logging.error(f"Error sending reminders: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============== AI ASSISTANT ==============

class AICustomerSummaryRequest(BaseModel):
    customer_id: str

@api_router.post("/ai/customer-summary")
async def get_ai_customer_summary(request: AICustomerSummaryRequest):
    """Generate AI-powered customer summary using Gemini"""
    try:
        if not GOOGLE_API_KEY:
            raise HTTPException(status_code=500, detail="Google API Key yapılandırılmamış")
        
        # Get customer data
        customer_response = supabase.table("customers").select("*").eq("id", request.customer_id).single().execute()
        if not customer_response.data:
            raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
        
        customer = customer_response.data
        
        # Get recent activities (calls, visits) - notes are stored in customer.notes_list
        calls_response = supabase.table("calls").select("*").eq("customer_id", request.customer_id).order("created_at", desc=True).limit(5).execute()
        visits_response = supabase.table("visits").select("*").eq("customer_id", request.customer_id).order("created_at", desc=True).limit(5).execute()
        
        # Notes are in customer.notes_list array
        notes_list = customer.get('notes_list', []) or []
        calls = calls_response.data or []
        visits = visits_response.data or []
        
        # Build context for AI
        context = f"""
Müşteri Bilgileri:
- Firma Adı: {customer.get('company_name', 'Bilinmiyor')}
- Market/Sektör: {customer.get('market', 'Belirtilmemiş')}
- Uygulama Alanı: {customer.get('application', 'Belirtilmemiş')}
- Şehir: {customer.get('city', 'Belirtilmemiş')}
- İlçe: {customer.get('district', 'Belirtilmemiş')}
- Web Sitesi: {customer.get('website', 'Yok')}
- Durum: {customer.get('status', 'Belirtilmemiş')}
- Potansiyel Seviyesi: {customer.get('potential_level', 'Belirtilmemiş')}
- Rakip: {customer.get('competitor', 'Belirtilmemiş')}
- Partner: {customer.get('partner', 'Belirtilmemiş')}
- Ürünler: {', '.join(customer.get('products', [])) if customer.get('products') else 'Belirtilmemiş'}
- Takip Eden: {customer.get('assigned_to', 'Belirtilmemiş')}
- Genel Notlar: {customer.get('notes', 'Yok')}

Son Notlar:
{chr(10).join([f"- {n.get('content', '')}" for n in notes_list[:3]]) if notes_list else 'Not yok'}

Son Aramalar:
{chr(10).join([f"- {c.get('outcome', 'Sonuç yok')}: {c.get('notes', '')}" for c in calls[:3]]) if calls else 'Arama kaydı yok'}

Son Ziyaretler:
{chr(10).join([f"- {v.get('visit_date', '')}: {v.get('notes', '')}" for v in visits[:3]]) if visits else 'Ziyaret kaydı yok'}
"""
        
        # System prompt for AI
        system_prompt = """Sen bir CRM asistanısın. Verilen müşteri bilgilerini analiz edip kısa ve öz bir Türkçe özet hazırlayacaksın.

Özet şu başlıkları içermeli:
1. 🏢 Firma Hakkında (1-2 cümle: Ne iş yapar, hangi sektörde)
2. 📊 Durum Özeti (1-2 cümle: Mevcut ilişki durumu)
3. 💡 Öneri (1 cümle: Bir sonraki adım ne olmalı)

Toplam 4-5 cümleyi geçme. Emoji kullan. Profesyonel ama samimi ol."""
        
        # Use Google Gemini API
        model = genai.GenerativeModel('gemini-2.5-flash')
        
        prompt = f"{system_prompt}\n\nBu müşteri hakkında kısa bir özet hazırla:\n\n{context}"
        
        response = await asyncio.to_thread(
            model.generate_content,
            prompt
        )
        
        summary_text = response.text if response.text else "Özet oluşturulamadı"
        
        return {
            "customer_id": request.customer_id,
            "company_name": customer.get('company_name'),
            "summary": summary_text,
            "generated_at": datetime.now(timezone.utc).isoformat()
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"AI Summary error: {e}")
        raise HTTPException(status_code=500, detail=f"AI özet oluşturulamadı: {str(e)}")

@api_router.post("/test-email")
async def send_test_email():
    """Send a test email to primary email address"""
    PRIMARY_EMAIL = "hakanonel05@gmail.com"
    try:
        html_content = """
        <!DOCTYPE html>
        <html>
        <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
            <div style="background: linear-gradient(135deg, #10b981, #059669); padding: 20px; border-radius: 10px;">
                <h1 style="color: white; margin: 0;">✅ CRMaster Test E-postası</h1>
            </div>
            <div style="background: #f8fafc; padding: 20px; border: 1px solid #e2e8f0; margin-top: 10px; border-radius: 10px;">
                <p style="color: #334155; font-size: 16px;">
                    Bu bir test e-postasıdır.<br><br>
                    E-posta bildirimleri başarıyla yapılandırıldı! 🎉
                </p>
                <p style="color: #64748b; font-size: 14px;">
                    CRMaster Müşteri Yönetim Sistemi
                </p>
            </div>
        </body>
        </html>
        """
        
        params = {
            "from": SENDER_EMAIL,
            "to": [PRIMARY_EMAIL],
            "subject": "✅ CRMaster - Test E-postası",
            "html": html_content
        }
        
        result = await asyncio.to_thread(resend.Emails.send, params)
        
        return {
            "message": f"Test e-postası {PRIMARY_EMAIL} adresine gönderildi",
            "email_id": result.get("id") if result else None
        }
    except Exception as e:
        logging.error(f"Test email error: {e}")
        raise HTTPException(status_code=500, detail=f"E-posta gönderilemedi: {str(e)}")

# Include router
app.include_router(api_router)

# CORS
cors_origins = os.environ.get('CORS_ORIGINS', '')
if cors_origins:
    origins_list = cors_origins.split(',')
else:
    origins_list = []

# Always include these origins
default_origins = [
    "http://localhost:3000",
    "http://127.0.0.1:3000",
    "https://button-consolidation.preview.emergentagent.com",
    "https://darling-dodol-27b576.netlify.app",
    "https://crmaster.netlify.app",
    "https://crmmaster-api.onrender.com"
]

for origin in default_origins:
    if origin not in origins_list:
        origins_list.append(origin)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=origins_list,
    allow_methods=["*"],
    allow_headers=["*"],
)

# GZip middleware - compresses JSON responses >500 bytes (70-90% smaller payloads)
app.add_middleware(GZipMiddleware, minimum_size=500, compresslevel=6)

# Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ============ STARTUP: BACKUP SCHEDULER ============
@app.on_event("startup")
def _start_backup_scheduler():
    try:
        backup_service.start_scheduler(_run_scheduled_backup)
    except Exception as e:
        logger.error("Failed to start backup scheduler: %s", e)


@app.on_event("shutdown")
def _stop_backup_scheduler():
    try:
        if backup_service._scheduler is not None:
            backup_service._scheduler.shutdown(wait=False)
    except Exception:
        pass
